from PyQt5.QtCore import *
from PyQt5.QtWidgets import *           ## importing classes of pyqt5 for applications
from PyQt5.QtGui import *
# Only needed for access to command line arguments
import sys
from pyqtgraph import PlotWidget, plot
import pyqtgraph as pg
#from dashboard import DashBoard
import openpyxl
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
#from dashboard import DashBoard
import matplotlib
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import pandas as pd
from rest_framework.response import Response
import json 
import requests
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from random import randint

import networkx as nx
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
class empty_window_for_loaders(QMainWindow):
    def __init__(self,path):
        super().__init__()
        self.setWindowFlag(Qt.FramelessWindowHint)
        self.resize(1000,640)
        self.setStyleSheet("background-color:white")
        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)
        self.setLayout(self.gridlayout)
        self.grid3 = QGridLayout(self)
        self.label = QLabel("None",self)
        self.label.setMaximumSize(600,400)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setLayout(self.grid3)

        self.movie = QMovie(path)
        self.label.setMovie(self.movie)
        self.label.setScaledContents(True)
        self.movie.setSpeed(100)
        self.movie.start()
        self.gridlayout.addWidget(self.label,0,0)
class empty_window_for_loaders2(QMainWindow):
    def __init__(self,path):
        super().__init__()
        self.setWindowFlag(Qt.FramelessWindowHint)
        self.resize(1000,640)
        self.setStyleSheet("background-color:white")
        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)
        self.setLayout(self.gridlayout)
        self.grid3 = QGridLayout(self)
        self.label = QLabel("None",self)
        self.label.setMaximumSize(600,400)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setLayout(self.grid3)
        self.label.setStyleSheet("background-color:black")

        self.movie = QMovie(path)
        self.label.setMovie(self.movie)
        self.label.setScaledContents(True)
        self.movie.setSpeed(100)
        self.movie.start()
        self.gridlayout.addWidget(self.label,0,0)

class RequestAndResponse():
    def __init__(self,endpoint):
        self.endpoint = endpoint
    def get_response(self):
        get_response = requests.get(self.endpoint)
        return get_response.json()

    def post_data(self,putdata):
        post = requests.post(self.endpoint, json=putdata)
        return post.text
    def delete_post(self,ids):
        for i in range(len(ids)):
            id = str(ids[i])
            updated_endpoint = self.endpoint + id 
            response = requests.delete(updated_endpoint)
    #if response = <Response [404]>:
        return response
    def delete_user(self,roll):
        response = requests.delete("http://127.0.0.1:8000/user/"+str(roll))
        return response
obj = RequestAndResponse("http://127.0.0.1:8000/login details/")
obj1 = RequestAndResponse("http://127.0.0.1:8000/user/")
print(obj.get_response())
class Ist_Ams(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlag(Qt.FramelessWindowHint)
        self.resize(1000,640)
        self.setStyleSheet( "QMainWindow{background-color:black}")
        self.setWindowIcon(QIcon('logo'))
    def Gui_login(self):
        self.widget = QWidget(self)
        self.layout = QGridLayout(self.widget)
        self.setCentralWidget(self.widget)
        
        self.label = QLabel("None",self)
        pixmap = QPixmap('loginbackground')
        self.label.setPixmap(pixmap)
        self.label.setScaledContents(True)
        self.label.setMaximumSize(1500,840)
        self.label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.label,0,0,3,3)

        self.ist_login_label = QLabel("𝗟𝗢𝗚𝗜𝗡",self)
        self.ist_login_label.setFont(QFont("arial",25))
        self.ist_login_label.setStyleSheet("text-decoration:underline")
        self.ist_login_label.setFixedSize(400,100)
        self.ist_login_label.setAlignment(Qt.AlignCenter)


        self.label =QLabel("login as",self)
        self.label.setFixedSize(100,40)
        self.label.setFont(QFont("arial",12))
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("background-color:rgba(255,255,255, 0.7);border:3px solid black;border-radius:5px;")
        self.combobox = QComboBox(self)
        self.combobox.setStyleSheet("background-color:rgba(255,255,255, 0.7);border:3px solid black;border-radius:5px;")
        self.combobox.setFont(QFont("arial",12))
        self.combobox.setMaximumSize(220,40)
        self.combobox.addItem("Faculty")
        self.combobox.addItem("Student")

        self.label2 =QLabel("Roll Number:",self)
        self.label2.setFont(QFont("arial",12))
        self.label2.setFixedSize(100,40)
        self.label2.setStyleSheet("background-color:rgba(255,255,255, 0.7);border:3px solid black;border-radius:5px;")
        self.roll_num = QLineEdit(self)
        self.roll_num.setStyleSheet("background-color:rgba(255,255,255, 0.7);border:3px solid black;border-radius:5px;")
        self.roll_num.setFont(QFont("arial",12))
        self.roll_num.setMaximumSize(220,40)

        self.label3 =QLabel("Password",self)
        self.label3.setFixedSize(100,40)
        self.label3.setFont(QFont("arial",12))
        self.label3.setStyleSheet("background-color:rgba(255,255,255, 0.7);border:3px solid black;border-radius:5px;")
        self.label3.setAlignment(Qt.AlignCenter)
        self.password = QLineEdit(self)
        self.password.setStyleSheet("background-color:rgba(255,255,255, 0.7);border:3px solid black;border-radius:5px;")
        self.password.setFont(QFont("arial",12))
        self.password.setMaximumSize(220,40)
        self.password.setEchoMode(QLineEdit.Password)
        
        ##nested layout## 

        self.button = QPushButton("ᴏᴋ",self)
        self.button.setFont(QFont("arial",15))
        self.button.setFixedSize(80,30)
        self.button.setStyleSheet('''QPushButton{background-color:white;color:black;border-radius:5px;border:2px solid black;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey;padding :2px solid black}
            ''')

        self.exit = QPushButton("Exit",self)
        self.exit.setFont(QFont("arial",15))
        self.exit.setFixedSize(80,30)
        self.exit.setStyleSheet('''QPushButton{background-color:white;color:black;border-radius:5px;border:2px solid black;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey;padding :2px solid black}
            ''')
        self.exit.clicked.connect(lambda:exitwindow())
        def exitwindow():
            stack.close()

        self.layout3 = QHBoxLayout(self)
        self.empty_label2 = QLabel(self)
        self.empty_label2.setFixedSize(200,50)
        self.empty_label2.setLayout(self.layout3)
        self.layout3.addWidget(self.button)
        self.layout3.addWidget(self.exit)

        self.layout2 = QFormLayout(self)
        self.empty_label =QLabel(self)
        self.empty_label.setAlignment(Qt.AlignBottom)
        self.empty_label.setMaximumSize(400,200)
        self.empty_label.setLayout(self.layout2)
        self.layout2.addWidget(self.label)
        self.layout2.addWidget(self.label2)
        self.layout2.addWidget(self.label3)
        self.layout2.addRow(self.label,self.combobox)
        self.layout2.addRow(self.label2,self.roll_num)
        self.layout2.addRow(self.label3,self.password)
        
        self.layout2.addWidget(self.empty_label2)
        self.layout.addWidget(self.ist_login_label,0,1,1,1)

        ##main window layout##
        self.layout.addWidget(self.empty_label,1,1,2,1)        
        self.button.clicked.connect(lambda:self.check())
        
    def check(self):
        login_details = obj.get_response()
        print("fed")
        print(login_details)
        for i in range(len(login_details)):
            #x = obj1.post({"user":self.line})
            if self.password.text() == (login_details[i]['password']) and self.combobox.currentText() == (login_details[i]['type']) and int(self.roll_num.text())==login_details[i]['rollno']:            
                print("sucess")
                print("jeafawf")
                self.adding_stack_elements()
                self.showloader()
            elif self.password.text() == (login_details[i]['password']) and self.combobox.currentText() == (login_details[i]['type']) and int(self.roll_num.text())==login_details[i]['rollno']:            
                print("sucess")
                self.dashboard = DashBoard()
                self.dashboard.Gui_dashboard()
                self.dashboard.show()
    def adding_stack_elements(self):
        #1
        self.empty = empty_window_for_loaders("loader1.gif")
        stack.addWidget(self.empty)
        #2
        self.dashboard = DashBoard()
        self.dashboard.Gui_dashboard()
        stack.addWidget(self.dashboard)
        #3
        self.faculty = Faculty()
        stack.addWidget(self.faculty)
        #4
        self.portal = Portal()
        stack.addWidget(self.portal)
        #5
        self.profile = Profile()
        stack.addWidget(self.profile)
        #6
        self.feedback = Feedback()
        stack.addWidget(self.feedback)
        x = stack.count()  
        #7
        self.department = Department()
        stack.addWidget(self.department)
        #8
        self.empty2 = empty_window_for_loaders2("logout1.gif")
        stack.addWidget(self.empty2)
        #9Events
        self.events = Events()
        stack.addWidget(self.events)
        #10Events
        self.msg = Messages()
        stack.addWidget(self.msg)        
    def returning(self):
        return [self.roll_num.text()]
    def showloader(self):
        stack.setCurrentIndex(1)
        QTimer.singleShot(4000,self.go_to_dashboard)
    def go_to_dashboard(self):
        stack.setCurrentIndex(2)
    def going_to_faculty_portal(self):
        stack.setCurrentIndex(3)
    def going_to_student_portal(self):
        stack.setCurrentIndex(4)
    def going_to_profile(self):
        stack.setCurrentIndex(5)
    def going_to_feedback(self): 
        stack.setCurrentIndex(6)
    def going_to_departent(self): 
        stack.setCurrentIndex(7)
    def logingout(self):
        stack.setCurrentIndex(8)
        QTimer.singleShot(7000,self.going_to_loign)
    def going_to_events(self):
        stack.setCurrentIndex(9)
    def going_to_messages(self):
        stack.setCurrentIndex(10)
    def going_to_loign(self):
        stack.deleteLater()
        self.obj = empty_window_for_loaders("loader3.gif")
        self.obj.show()
        status = QProcess.startDetached(sys.executable, sys.argv)
        QTimer.singleShot(8000,self.obj.close)
    def going_back_to_dashboard(self):
        stack.setCurrentIndex(2)


class DashBoard(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setStyleSheet("QMainWindow{background-image:url(dash_background)}")
    def Gui_dashboard(self):
        self.widget = QWidget(self)  # dummy widget to add the layout too
        self.setCentralWidget(self.widget) 
        self.layout = QGridLayout(self.widget)
        
        self.istlabel = QLabel(self)
        self.istlabel.setPixmap(QPixmap("ist-img"))
        self.istlabel.setMaximumWidth(350)
        self.istlabel.setMaximumHeight(250)
        getdata = RequestAndResponse("http://127.0.0.1:8000/quiz marks/")
        result = getdata.get_response()
        print(result)
        maximumnumber,obtainnumber = [],[]
        calculation = [0,0,0,0,0,0,0,0,0,0]
        for i in range (len(result)):
            if result[i]['typeofexam'] == "Quiz":
                if result[i]['rollno'] == 210201073:
                    print("huoafdhoz")
                    maximumnumber.append(result[i]['maxnum'])
                    obtainnumber.append(result[i]['obtainnum'])
        for i in range (len(obtainnumber)):     
                print(maximumnumber,obtainnumber)
                calculation[i] = obtainnumber[i]/maximumnumber[i]*100
        print(calculation)
        self.plot = pg.plot()
        self.plot
        y1 = calculation
        xlab = ['quiz1', 'quiz2', 'quiz3', 'quiz4', 'asg1', 'asg2','proj', 'oht1', 'oht2', 'final']
        xval = list(range(1,len(xlab)+1))
        ticks=[]
        for i, item in enumerate(xlab):
            ticks.append( (xval[i], item) )
        ticks = [ticks]


        self.bargraph = pg.BarGraphItem(x = xval, height = y1, width=0.5, brush ='lightblue')
        self.plot.addItem(self.bargraph)

        ax = self.plot.getAxis('bottom')
        ax.setTicks(ticks)
        self.plot.setYRange(0,100)
        self.plot.setBackground("white")
        self.plot.setMaximumSize(700,450)
        self.plot.setStyleSheet("border:2px solid black;border-radius:5%")
        self.hv = QHBoxLayout(self)
        self.vb = QVBoxLayout(self)
        self.horizontal_streach = QLabel(self)
        self.horizontal_streach.setLayout(self.hv)
        self.horizontal_streach.setLayout(self.vb)
        self.hv.addWidget(self.plot)

        if window_1.returning()[0] == "0":
            self.horizontal_streach.hide()
            def plot(color,color2,color3):
                getdata = RequestAndResponse("http://127.0.0.1:8000/ user_score/")
                result = getdata.get_response()
                print(result)
                val = result[0]['user_score']
                self.figure = plt.figure(facecolor=color3)
                self.canvas = FigureCanvas(self.figure)
                ax = self.figure.add_subplot(111)
                plt.ylim((0,100))
                x = ["-.",".-", "your obtained percentage by students feedbacks", "-","."]
                h = [0, 0, int(val), 0, 0]
                ax.bar(x, h,color=color,width=0.4)
                ax.set_facecolor(color2)
                self.canvas.setMaximumSize(700,450)
                self.hlayout = QHBoxLayout(self)
                self.vlayout = QVBoxLayout(self)
                self.emptye = QLabel(self)
                self.emptye.setStyleSheet("border:3px solid black;background-color:white")
                self.emptye.setLayout(self.hlayout)
                self.emptye.setLayout(self.vlayout)
                self.hlayout.addWidget(self.canvas)
                self.layout.addWidget(self.emptye,2,5,3,4)
            plot("lightblue","white","white")


        
        self.Side_label = QLabel("",self)
    #   self.Side_label.move(0,0)
        self.Side_label.setMaximumSize(200,1000)
        self.Side_label.setFont(QFont("arial",20))
        self.Side_label.setStyleSheet("background-color:rgb(255, 255, 255,0.4);border-radius:10px;border:2px solid black;")
        self.Side_label_layout = QGridLayout(self)
        self.Side_label.setLayout(self.Side_label_layout)



        self.darkmode_button = QPushButton("Dark Mode",self)
        self.darkmode_button.setMaximumSize(150,40)
        self.darkmode_button.setStyleSheet('''QPushButton{background-color:black;color:white;}
        QPushButton:hover{border:2px solid white;}
        QPushButton:pressed{background-color:white;color:black;padding:5px solid black;}''')
        def change_color():
            try:
                plot("lightgrey","grey","grey")
                self.emptye.setStyleSheet("border:3px solid black;background-color:black")

            except:
                print("faculty porta ha ya lanti")
            self.setStyleSheet("QMainWindow{background-image:url(darkbackground2)}")
            self.button3.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:10px solid black;}''')
            self.bargraph.setOpacity(0.3)
            self.plot.setBackground("black")
            self.M_button.setStyleSheet('''QPushButton{background-color:grey;border-radius:10px;border:2px solid black;}
            QPushButton:hover{background-color:white;}
            QPushButton:pressed{padding:5px solid black;}''')       
            self.u_button.setStyleSheet('''QPushButton{background-color:grey;border-radius:10px;border:2px solid black;}
            QPushButton:hover{background-color:white;}
            QPushButton:pressed{padding:5px solid black;}''')
            self.button.setStyleSheet('''QPushButton{background-color:grey;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:10px solid black;}''')
            self.button2.setStyleSheet('''QPushButton{background-color:grey;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
            self.button3.setStyleSheet('''QPushButton{background-color:grey;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
        
        self.darkmode_button.clicked.connect(lambda:change_color())

        self.lightmode_button = QPushButton("Light Mode",self)
        self.lightmode_button.setMaximumSize(150,40)
        self.lightmode_button.setStyleSheet('''QPushButton{background-color:rgb(255, 255, 255,0.4);color:black;}
        QPushButton:hover{border:3px solid black;background-color:white}
        QPushButton:pressed{background-color:lightgrey;padding:5px solid black;}''')
        def change_color2():
            try:
                plot("lightblue","white","white")
            except:
                print("faculty porta ha ya lanti")
            self.setStyleSheet("QMainWindow{background-image:url(dash_background)}")
            self.istlabel.show()
            self.button3.setStyleSheet('''QPushButton{background-color: srgb(135,206,235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:10px solid black;}''')
            self.bargraph.setOpacity(1)
            self.plot.setBackground("white")
            self.M_button.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
            QPushButton:hover{background-color:white;}
            QPushButton:pressed{padding:5px solid black;}''')       
            self.u_button.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
            QPushButton:hover{background-color:white;}
            QPushButton:pressed{padding:5px solid black;}''')
            self.button.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:10px solid black;}''')
            self.button2.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
            self.button3.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
        self.logout_button = QPushButton("Logout",self)
        self.logout_button.setMaximumSize(180,40)
        self.logout_button.setFont(QFont("arial",15))
        self.logout_button.setStyleSheet('''
        QPushButton{border:3px solid black;}
        QPushButton:hover{border:3px solid red;background-color:white;}
        QPushButton:pressed{background-color:lightgrey;padding:5px solid black;}''')
        self.logout_button.clicked.connect(window_1.logingout)

        self.p_button = QPushButton("",self)
        self.p_button.setFont(QFont("arial",20))
        self.p_button.setStyleSheet('''QPushButton{background-image:url(profilecopy);border-radius:10px;border:2px solid black}
        QPushButton:hover{border:2px solid white}
        QPushButton:pressed{border:2px solid transparent}''')
        self.p_button.setMaximumWidth(180)
        self.p_button.setMaximumHeight(200)
        #self.p_button.setAlignment(Qt.AlignHCenter)
        if window_1.returning()[0] == "210201073":
            self.p_button.clicked.connect(lambda:window_1.going_to_profile())

        self.lightmode_button.clicked.connect(lambda:change_color2())
        self.modes = QVBoxLayout(self)
        self.modes_empty_label = QLabel(self)
        self.modes_empty_label.setStyleSheet("background-color:rgb(255, 255, 255,0.4)")
        self.modes_empty_label.setMaximumSize(170,100)
        self.modes_empty_label.setLayout(self.modes)
        self.modes.addWidget(self.darkmode_button)
        self.modes.addWidget(self.lightmode_button)     
        self.empty_labele = QLabel(self)
        self.empty_labele.setStyleSheet("background-color:transparent;border:2px solid transparent")
        self.Side_label_layout.addWidget(self.modes_empty_label,0,0,1,1)
        self.Side_label_layout.addWidget(self.p_button,2,0,1,1)
        self.Side_label_layout.addWidget(self.empty_labele,1,0,1,1)
        self.Side_label_layout.addWidget(self.logout_button,3,0,1,1)
        



        self.M_label = QLabel("",self)
    #   self.M_label.move(1650,0)
    #   self.M_label.resize(250,50)
        self.M_label.setFont(QFont("arial",20))
        self.M_label.setStyleSheet('''QLabel{background-image:url(messages);}''')
        self.M_label.setMaximumWidth(400)
        self.M_label.setMaximumHeight(350)
        self.M_button = QPushButton("<Messages>",self)
        self.M_button.setFont(QFont("arial",20))
        self.M_button.setStyleSheet("")
        self.M_button.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
            QPushButton:hover{background-color:white;}
            QPushButton:pressed{padding:5px solid black;}''')     
        self.M_button.clicked.connect(window_1.going_to_messages)  
        self.M_button.setMaximumWidth(400)
        self.M_button.setMaximumHeight(50)
        self.M_label.setScaledContents(True)
        self.layout_for_messages = QVBoxLayout(self)
        self.layout_for_messages.addWidget(self.M_label)
        self.layout_for_messages.addWidget(self.M_button)
        self.empty_label_for_messages = QLabel(self)
        self.empty_label_for_messages.setLayout(self.layout_for_messages)
        self.empty_label_for_messages.setStyleSheet("background-color:transparent;border-radius:5px solid black;")
        
        self.u_button = QPushButton("<Events>  ",self)
        self.u_button.setFont(QFont("arial",20))
        self.u_button.setStyleSheet('''QPushButton{background-color:rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
            QPushButton:hover{background-color:white;}
            QPushButton:pressed{padding:5px solid black;}''')
        self.u_button.setMaximumWidth(400)
        self.u_button.setMaximumHeight(50)
        self.u_button.clicked.connect(window_1.going_to_events)
        self.u_label = QLabel("",self)
    #   self.u_button.move(1250, 50)
    #   self.u_button.resize(400,350)
        self.u_label.setFont(QFont("arial",20))
        self.u_label.setStyleSheet('''QLabel{border-radius:10px;border-radius:5px solid black;background-image:url(eventsbutton);color:white;}''')      
        self.u_label.setMaximumWidth(400)
        self.u_label.setMaximumHeight(350)
        self.u_label.setScaledContents(True)
        self.layout_for_events = QVBoxLayout(self)
        self.layout_for_events.addWidget(self.u_label)
        self.layout_for_events.addWidget(self.u_button)
        
        self.empty_label = QLabel(self)
        self.empty_label.setLayout(self.layout_for_events)
        self.empty_label.setStyleSheet("background-color:transparent;border-radius:5px solid black")

        self.button = QPushButton(" <IST PORTAL>",self)
    #   self.button.setFixedSize(600,100)
        self.button.setStyleSheet("")
        self.button.setStyleSheet('''QPushButton{background-color: rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
    #   self.button.move(300,525)           
        self.button.setFont(QFont("arial",15))
        self.button.setMaximumSize(250,100)
        if window_1.returning()[0] == "210201073":
            self.button.clicked.connect(lambda:window_1.going_to_student_portal())
        else:
            self.button.clicked.connect(lambda:window_1.going_to_faculty_portal())
        


        self.button2 = QPushButton("  <FEEDBACK>",self)
    #   self.button2.resize(600,100)
        self.button2.setStyleSheet("")
        self.button2.setStyleSheet('''QPushButton{background-color: rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
    #   self.button2.move(300,650)           
        self.button2.setFont(QFont("arial",15))
        self.button2.setMaximumSize(250,100)
        if window_1.returning()[0] == "210201073":
            self.button2.clicked.connect(window_1.going_to_feedback)

        self.button3 = QPushButton("  <Department> ",self)
    #   self.button3.resize(600,100)
        self.button3.setStyleSheet("")
        self.button3.setStyleSheet('''QPushButton{background-color: rgb(135, 206, 235,0.7);border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding:5px solid black;}''')
    #   self.button3.move(300,775)           
        self.button3.setFont(QFont("arial",15))
        self.button3.setMaximumSize(250,100)
        self.button3.clicked.connect(window_1.going_to_departent)
        
        self.layout_three_button = QVBoxLayout(self)
        self.layout_three_button.addWidget(self.button)
        self.layout_three_button.addWidget(self.button2)
        self.layout_three_button.addWidget(self.button3)
        self.empty_label_for_three_buttons = QLabel(self)
        self.empty_label_for_three_buttons.setMaximumSize(300,200)
        self.empty_label_for_three_buttons.setLayout(self.layout_three_button)
        self.empty_label_for_three_buttons.setStyleSheet("background-color:transparent;border-radius:5px solid black")

        self.class_reputation = QLabel("")

        self.layout.addWidget(self.Side_label,0,0,5,2)
        self.layout.addWidget(self.istlabel,0,3,2,2)
        self.layout.addWidget(self.empty_label,0,5,2,2)
        self.layout.addWidget(self.empty_label_for_messages,0,7,2,2)
        self.layout.addWidget(self.empty_label_for_three_buttons,3,3,2,2)
        self.layout.addWidget(self.horizontal_streach,2,5,3,4)
        self.resize(1000,640)


           


class Portal(QMainWindow):
    def __init__(self):
        super().__init__()
        w=1000
        h=640
        self.setStyleSheet( '''QMainWindow{background-color:black}QScrollBar{ background-color: darkgrey}QTableWidget::item {background-color:lightblue;border: 1px solid black;color:black}
            QTableWidget::item::hover{background-color:white}QTableWidget::item::selected{background-color:white}
            QHeaderView::section {border:1px solid black;}QHeaderView:section:hover{background-color:lightblue}''')

        self.setWindowIcon(QIcon('logo'))
        self.resize(w,h)

        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)

        self.grid3 = QGridLayout(self)
        self.label = QLabel("None",self)
        self.label.setStyleSheet("background-color:black;")
        self.label.setMaximumSize(2000,840)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setLayout(self.grid3)

        self.movie = QMovie("loader.gif")
        self.label.setMovie(self.movie)
        self.label.setScaledContents(True)
        self.movie.setSpeed(200)
        self.movie.start()

        self.hb = QHBoxLayout(self)
        self.empty_label2 = QLabel(self)
        self.empty_label2.setFixedSize(170,50)
        self.empty_label2.setStyleSheet("border:2px solid transparent")
        self.empty_label2.setLayout(self.hb)
        self.student_label = QLabel("Students",self)
        self.student_label.setStyleSheet("background-color:transparent;color:rgba(72,221,253,255);border-style:none")
        self.student_label.setMaximumSize(150,30)
        self.student_label.setAlignment(Qt.AlignCenter)
        self.student_label.setFont(QFont("arial",15))


        self.goback = QPushButton("<",self)
        self.goback.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.goback.setFixedSize(30,30)
        self.hb.addWidget(self.goback)
        self.hb.addWidget(self.student_label)
        self.goback.clicked.connect(lambda:window_1.going_back_to_dashboard())


        self.attendance = QPushButton("Attendance",self)
        self.attendance.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{color:white;border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.attendance.setFixedSize(170,30)
        

        self.assembly_language = QAction("computer organization and assembly lamguage")
        self.assembly_language_lab = QAction("computer organization and assembly lamguage lab")
        self.data_structure = QAction("data structure")
        self.data_structure_lab = QAction("data structure lab")
        self.discrete_structure = QAction("discrete structure")
        self.graph_theory = QAction("graph theory")
        self.profesional_practices = QAction("professional practices",print("profesional_practices"))       
        self.menu = QMenu()

        self.menu.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255);}
            QMenu:pressed{background-color:lightblue}''')
        menu = self.menu.addMenu("courses")
        menu.addAction(self.assembly_language)
        menu.addAction(self.assembly_language_lab)
        menu.addAction(self.data_structure)
        menu.addAction(self.data_structure_lab)
        menu.addAction(self.graph_theory)
        menu.addAction(self.profesional_practices)
        menu.addAction(self.assembly_language_lab)
        self.attendance.setMenu(self.menu)


        self.result = QPushButton("results",self)
        self.result.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.result.setFixedSize(170,30)
        
        self.quizes = QAction("Quizes")
        self.assignments = QAction("Assignments")
        self.project = QAction("Project")
        self.presentation = QAction("presentation")
        self.oht1 = QAction("oht1")
        self.oht2 = QAction("oht2")
        self.final = QAction("final")
        self.menu_for_results = QMenu()
        self.menu_for_results.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255)}
            QMenu:pressed{background-color:lightblue}''')

        self.sub_menu_quiz = self.menu_for_results.addMenu("Quizes")
        self.sub_menu_assignment = self.menu_for_results.addMenu("Assignments")
        self.sub_menu_project = self.menu_for_results.addMenu("Project")
        self.sub_menu_presentation = self.menu_for_results.addMenu("Presentation")
        self.sub_menu_oht1 = self.menu_for_results.addMenu("OHT 1")
        self.sub_menu_oht2 = self.menu_for_results.addMenu("OHT 2")
        self.sub_menu_final = self.menu_for_results.addMenu("Final")
        self.result.setMenu(self.menu_for_results)

        self.sub_menu_quiz.addAction("computer organization and assembly language",lambda:print("hello"))
        self.sub_menu_quiz.addAction("computer organization and assembly language")
        self.sub_menu_quiz.addAction("data structure",lambda:self.marks_table())
        self.sub_menu_quiz.addAction("data structure lab")
        self.sub_menu_quiz.addAction("graph theory")
        self.sub_menu_quiz.addAction("professional practices")

        self.sub_menu_assignment.addAction("computer organization and assembly language",lambda:print("hello"))
        self.sub_menu_assignment.addAction("computer organization and assembly language")
        self.sub_menu_assignment.addAction("data structure")
        self.sub_menu_assignment.addAction("data structure lab")
        self.sub_menu_assignment.addAction("graph theory")
        self.sub_menu_assignment.addAction("professional practices")


        self.sub_menu_project.addAction("computer organization and assembly language",lambda:print("hello"))
        self.sub_menu_project.addAction("computer organization and assembly language")
        self.sub_menu_project.addAction("data structure")
        self.sub_menu_project.addAction("data structure lab")
        self.sub_menu_project.addAction("graph theory")
        self.sub_menu_project.addAction("professional practices")


        self.sub_menu_presentation.addAction("computer organization and assembly language",lambda:print("hello"))
        self.sub_menu_presentation.addAction("computer organization and assembly language")
        self.sub_menu_presentation.addAction("data structure")
        self.sub_menu_presentation.addAction("data structure lab")
        self.sub_menu_presentation.addAction("graph theory")
        self.sub_menu_presentation.addAction("professional practices")

        self.sub_menu_oht1.addAction("computer organization and assembly language",lambda:print("hello"))
        self.sub_menu_oht1.addAction("computer organization and assembly language")
        self.sub_menu_oht1.addAction("data structure")
        self.sub_menu_oht1.addAction("data structure lab")
        self.sub_menu_oht1.addAction("graph theory")
        self.sub_menu_oht1.addAction("professional practices")

        self.sub_menu_oht2.addAction("computer organization and assembly language")
        self.sub_menu_oht2.addAction("computer organization and assembly language")
        self.sub_menu_oht2.addAction("data structure")
        self.sub_menu_oht2.addAction("data structure lab")
        self.sub_menu_oht2.addAction("graph theory")
        self.sub_menu_oht2.addAction("professional practices")


        self.sub_menu_final.addAction("computer organization and assembly language")
        self.sub_menu_final.addAction("computer organization and assembly language")
        self.sub_menu_final.addAction("data structure")
        self.sub_menu_final.addAction("data structure lab")
        self.sub_menu_final.addAction("graph theory")
        self.sub_menu_final.addAction("professional practices")
        
        self.route = QPushButton("route",self)
        self.route.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{color:white;border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.route.setFixedSize(170,30)
        self.route.clicked.connect(self.transport)


        self.formlayout = QVBoxLayout(self)
        self.empty_label = QLabel(self)
        self.empty_label.setLayout(self.formlayout)
        self.empty_label.setFixedSize(220,700)
        self.empty_label.setStyleSheet("border:10px solid #52bdf6;border-style:double;border-right:transparent")
        self.formlayout.addWidget(self.empty_label2)
        self.formlayout.addWidget(self.attendance)
        self.formlayout.addWidget(self.result)
        self.formlayout.addWidget(self.route)
        self.formlayout.addStretch(1)
        self.scrolle = QScrollArea(self)
        self.scrolle.setWidget(self.empty_label)
        self.scrolle.setWidgetResizable(True)
        self.scrolle.setMaximumSize(220,820)
        self.scrolle.setStyleSheet("border:2px solid #52bdf6;background-color:transparent")

        self.data_structure.triggered.connect(lambda:self.attendance_table())
        
        self.gridlayout.addWidget(self.label,0,0,3,3)
        self.gridlayout.addWidget(self.scrolle,0,0,3,1)
    
    def attendance_table(self):
        self.attendance.setText("assembly language..")
        self.tableWidget = QTableWidget(self)
        self.tableWidget.verticalHeader().setVisible(False) 
        self.gridlayout.addWidget(self.tableWidget,1,1,2,3)
        self.tableWidget.setMaximumSize(1500,800)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        self.layout_for_attendance_edit = QHBoxLayout(self)

        self.attendance_label = QLabel("Your Attendance",self)
        self.attendance_label.setStyleSheet("color:black;background-color:white")
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.attendance_label.setMaximumSize(1500,50)
        self.gridlayout.addWidget(self.attendance_label,0,1,1,3)
        getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
        result = getdata.get_response()
        print(result)
        ids,date,name,rollno,status= [],[],[],[],[]
        for i in range (len(result)):
            if result[i]['regno'] == 210201073:
                ids.append(result[i]['id'])
                date.append(result[i]['date'])
                name.append(result[i]['name'])
                rollno.append(result[i]['regno'])
                status.append(result[i]['status'])
        maxrow = 50
        maxcol = 5
        self.tableWidget.setRowCount(maxrow) 
        self.tableWidget.setColumnCount(maxcol)  
        currentRowCount = -1
        for i in range(len(ids)):
            currentRowCount += 1
            print("currentRowCount",currentRowCount)
            self.tableWidget.setItem(currentRowCount, 0, QTableWidgetItem(str(i+1)))
            self.tableWidget.setItem(currentRowCount, 1, QTableWidgetItem(str(date[i])))
            self.tableWidget.setItem(currentRowCount, 2, QTableWidgetItem(str(name[i])))
            self.tableWidget.setItem(currentRowCount, 3, QTableWidgetItem(str(rollno[i])))
            self.tableWidget.setItem(currentRowCount, 4, QTableWidgetItem(str(status[i])))


        
        def QString(a):
            return (a)
        self.tableWidget.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))

            

    def marks_table(self):
        self.markstable = QTableWidget(self)
        self.markstable.horizontalHeader().setStretchLastSection(True)
        self.markstable.verticalHeader().setVisible(False)  
        self.gridlayout.addWidget(self.markstable,1,1,2,3)
        self.markstable.setMaximumSize(1500,800)
        self.markstable.verticalHeader().setVisible(False)
        self.markstable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.markstable.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.layout_for_attendance_edit = QHBoxLayout(self)

        self.attendance_label = QLabel("Your Marks",self)
        self.attendance_label.setStyleSheet("color:black;background-color:white")
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.attendance_label.setMaximumSize(1500,50)
        self.gridlayout.addWidget(self.attendance_label,0,1,1,3)
        getdata = RequestAndResponse("http://127.0.0.1:8000/quiz marks/")
        result = getdata.get_response()
        print(result)
        ids,date,name,rollno,obtainnumber,maximumnumber= [],[],[],[],[],[]
        for i in range (len(result)):
            if result[i]['typeofexam'] == "Quiz":
                ids.append(result[i]['id'])
                date.append(result[i]['date'])
                name.append(result[i]['name'])
                rollno.append(result[i]['rollno'])
                maximumnumber.append(result[i]['maxnum'])
                obtainnumber.append(result[i]['obtainnum'])

        maxrow = 50
        maxcol = 6
        self.markstable.setRowCount(maxrow) 
        self.markstable.setColumnCount(maxcol)  
        currentRowCount = -1
        for i in range(len(ids)):
            currentRowCount += 1
            print("currentRowCount",currentRowCount)
            self.markstable.setItem(currentRowCount, 0, QTableWidgetItem(str(i+1)))
            self.markstable.setItem(currentRowCount, 1, QTableWidgetItem(str(date[i])))
            self.markstable.setItem(currentRowCount, 2, QTableWidgetItem(str(name[i])))
            self.markstable.setItem(currentRowCount, 3, QTableWidgetItem(str(rollno[i])))
            self.markstable.setItem(currentRowCount, 4, QTableWidgetItem(str(maximumnumber[i])))
            self.markstable.setItem(currentRowCount, 5, QTableWidgetItem(str(obtainnumber[i])))

        
        
        def QString(a):
            return (a)
        self.markstable.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Maximum Marks;Obtain").split(";"))
    def transport(self):
        self.emptylabel = QLabel(self) 
        self.gridlayout.addWidget(self.emptylabel,0,1,3,3)
        self.emptylabel.setMaximumSize(1500,800)
        self.emptylabel.setStyleSheet("background-color:white")
    
        self.newgrid = QGridLayout(self)
        self.emptylabel.setLayout(self.newgrid)

        self.combobx = QComboBox(self)
        list_of_cities = ["from","islamabad blue area","rawalpindi","Wah cantt","Westridge","scheme3","faizabad"]
        self.combobx.addItems(list_of_cities)
        self.combobx.setMaximumSize(300,40)
        self.combobx.setStyleSheet("background-color:lightblue")

        self.isb = QLabel("Islamabad | T-chowk",self)
        self.isb.setFont(QFont("arial", 15))
        self.isb.setAlignment(Qt.AlignCenter)
        self.isb.setStyleSheet("background-color:lightblue")
        self.to = QLabel("TO",self)
        self.to.setStyleSheet("background-color:lightblue")
        self.to.setFont(QFont("arial", 15))
        self.to.setAlignment(Qt.AlignCenter)

        self.hv2 = QHBoxLayout(self) 
        self.emptylabel3 = QLabel(self)
        self.emptylabel3.setLayout(self.hv2)
        self.path = QLabel("Shortest PATH:",self)
        self.path.setFont(QFont("arial",15))
        self.path.setAlignment(Qt.AlignCenter)
        self.path.setMaximumSize(200,100)

        self.result_path = QLabel("")
        self.result_path.setFont(QFont("arial",15))
        self.result_path.setWordWrap(True)
        
        self.hv2.addWidget(self.path)
        self.hv2.addWidget(self.result_path)
        self.figure = plt.figure()
        self.figure.clear()
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setMaximumSize(1500,450)
        self.canvas.setStyleSheet("background-color:black")
        self.canvas.hide()
        self.newgrid.addWidget(self.combobx,0,0)
        self.newgrid.addWidget(self.to,0,1)
        self.newgrid.addWidget(self.isb,0,2)
        self.newgrid.addWidget(self.canvas,2,0,1,3)
        self.newgrid.addWidget(self.emptylabel3,3,0,2,3)
        def showmore():
            self.canvas.show()
            graph_list = [["I S T","rawalpindi",15],["I S T","islamabad blue area",20],["rawalpindi","Westridge",15],
            ["rawalpindi","scheme3",1],["islamabad blue area","faizabad",10],["islamabad blue area","Wah cantt",40],["I S T","faizabad",30],["faizabad","Wah cantt",10]]
            graph = nx.Graph()
            graph.add_weighted_edges_from(graph_list)
            try:
                path = nx.dijkstra_path(graph,source="I S T", target=str(self.combobx.currentText()))
                self.result_path.setText(str(path))
                plt.rcParams['axes.facecolor'] = 'black'    
                graph = nx.Graph()
                graph.add_weighted_edges_from(graph_list)
                pos = nx.planar_layout(graph)
                nx.draw(graph, pos, with_labels=True,node_color='lightblue',node_size=700)
                #nx.to_undirected(graph,as_view=False)
                nx.draw_networkx_edge_labels(graph,
                                     pos,
                                 edge_labels={(u, v): d for u, v, d in graph.edges(data="weight")},
                                 )
                print(graph_list)
            except:
                print("dont add from")
        self.combobx.currentIndexChanged.connect(lambda:showmore())
class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        # last column
        if index.column() == (index.model().columnCount() - 1 ):
            return super().createEditor(parent, option, index)

        if index.column() == (index.model().columnCount() - 2 ):
            return super().createEditor(parent, option, index)



class Faculty(QMainWindow):
    date = None
    maxmarks = None
    def __init__(self):
        super().__init__()
        w=1000
        h=640
        self.resize(w,h)
        self.setStyleSheet( '''QMainWindow{background-color:black}QScrollBar{ background-color: darkgrey}QTableWidget::item {background-color:lightblue;border: 1px solid black;color:black}
            QTableWidget::item::hover{background-color:white}QTableWidget::item::selected{background-color:white}
            QHeaderView::section {border:1px solid black;}QHeaderView:section:hover{background-color:lightblue}''')
        self.setWindowIcon(QIcon('logo'))

        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)
        self.setLayout(self.gridlayout)

        self.grid3 = QGridLayout(self)
        self.label = QLabel("None",self)
        self.label.setStyleSheet("background-color:black;")
        self.label.setMaximumSize(1500,800)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setLayout(self.grid3)

        self.movie = QMovie("loader.gif")
        self.label.setMovie(self.movie)
        self.label.setScaledContents(True)
        self.movie.setSpeed(200)
        self.movie.start()

        self.hb = QHBoxLayout(self)
        self.empty_label2 = QLabel(self)
        self.empty_label2.setFixedSize(170,50)
        self.empty_label2.setStyleSheet("border:2px solid transparent")
        self.empty_label2.setLayout(self.hb)
        self.Faculty_label = QLabel(" Faculty",self)
        self.Faculty_label.setStyleSheet("background-color:transparent;color:rgba(72,221,253,255);border-style:none")
        self.Faculty_label.setFixedSize(150,30)
        self.Faculty_label.setFont(QFont("arial",15))
        self.goback = QPushButton("<",self)
        self.goback.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.goback.setFixedSize(30,30)
        self.goback.clicked.connect(lambda:window_1.going_back_to_dashboard())
        self.hb.addWidget(self.goback)
        self.hb.addWidget(self.Faculty_label)


        self.add_attendance = QPushButton("Add Attendance",self)
        self.add_attendance.setFixedSize(150,30)
        self.add_attendance.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.custom = QAction("Add Attendance Customly", self)
        self.upload_excel= QAction("Upload Excel File", self)
        menu3 = QMenu()
        menu3.addAction(self.custom)
        menu3.addAction(self.upload_excel)
        self.add_attendance.setMenu(menu3)
        menu3.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255)}
            QMenu:pressed{background-color:lightblue}''')
        self.custom.triggered.connect(lambda:self.showeditabletable())
        self.upload_excel.triggered.connect(lambda:self.upload_excelfile())
        self.attendance_information = QPushButton("Attendance information",self)
        self.attendance_information.setStyleSheet("background-color:black;color:rgba(72,221,253,255);border:2px solid rgba(72,221,253,255);border-radius :7px;border-radius :7px;")
        self.attendance_information.setFixedSize(150,30)
        self.attendance_information.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.assembly_language = QAction("Computer Organization And Assembly Language")
        self.assembly_language_lab = QAction("Computer Organization And Assembly Language lab")
        self.data_structure = QAction("Data Structure")
        self.data_structure_lab = QAction("Data Structure Lab")
        self.discrete_structure = QAction("Discrete Structure")
        self.graph_theory = QAction("Graph Theory")
        self.profesional_practices = QAction("Professional Practices",print("Profesional_Practices"))       
        self.menu = QMenu()
        menu1 = self.menu.addMenu("Courses")
        menu1.addAction(self.assembly_language)
        menu1.addAction(self.assembly_language_lab)
        menu1.addAction(self.data_structure)
        menu1.addAction(self.data_structure_lab)
        menu1.addAction(self.graph_theory)
        menu1.addAction(self.profesional_practices)
        menu1.addAction(self.assembly_language_lab)
        self.attendance_information.setMenu(self.menu)
        self.menu.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255)}
            QMenu:pressed{background-color:lightblue}''')
        self.data_structure.triggered.connect(lambda:self.showtable()) 
        

        self.result = QPushButton("Marks Upload",self)
        self.result.setFixedSize(150,30)
        self.result.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.result.clicked.connect(lambda:self.FacultyMarks())

        self.grade_criteria = QPushButton("Set Marks Distribution",self)
        self.grade_criteria.setFixedSize(150,30)
        self.grade_criteria.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}''')
        self.grade_criteria.clicked.connect(lambda:self.showgradingcriteria())

        self.grade_generation = QPushButton("Grade Generation",self)
        self.grade_generation.setFixedSize(150,30)
        self.grade_generation.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}''')
        self.grade_generation.clicked.connect(self.gradesMaker)

        self.form_for_buttons = QFormLayout(self)
        self.empty_label_for_buttons = QLabel(self)
        self.empty_label_for_buttons.setStyleSheet("border:10px solid #52bdf6;border-style:double;border-right:transparent")
        self.empty_label_for_buttons.setFixedSize(220,700)
        self.empty_label_for_buttons.setLayout(self.form_for_buttons)
        self.form_for_buttons.addWidget(self.empty_label2)
        self.form_for_buttons.addWidget(self.add_attendance)
        self.form_for_buttons.addWidget(self.attendance_information)
        self.form_for_buttons.addWidget(self.result)
        self.form_for_buttons.addWidget(self.grade_criteria)
        self.form_for_buttons.addWidget(self.grade_generation)
        
        self.scroll = QScrollArea(self)
        self.scroll.setWidget(self.empty_label_for_buttons)
        self.scroll.setWidgetResizable(True)
        self.scroll.setMaximumSize(220,800)
        self.scroll.setStyleSheet("border:2px solid #52bdf6;background-color:transparent")

        self.gridlayout.addWidget(self.label,0,0,5,5)         
        self.gridlayout.addWidget(self.scroll,0,0,5,1)        

    def showeditabletable(self):
        self.empty_label1 = QLabel(self)
        self.empty_label1.setStyleSheet("color:black;background-color:white")
        self.empty_label1.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.upload = QPushButton("Upload",self)
        self.upload.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.upload.setFixedSize(120,30)
        self.upload.clicked.connect(lambda:upload())
        self.attendance_label = QLabel("Students Attendance",self)
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label1.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.upload)

        self.editable_tableWidget = QTableWidget(40,5,self)
        self.editable_tableWidget.setMaximumSize(1500,800)        
        self.editable_tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        maxrow = 50
        maxcol = 5
        self.editable_tableWidget.setRowCount(maxrow) 
        self.editable_tableWidget.setColumnCount(maxcol)  


        def QString(a):
            return (a)
        self.editable_tableWidget.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))
        self.gridlayout.addWidget(self.empty_label1,0,1,1,5)
        self.gridlayout.addWidget(self.editable_tableWidget,1,1,4,5)
        def upload():
            whole = {}
            x=""
            counter = -1
            for i in range(maxrow):
                individual_dict = {}
                for j in range(maxcol):

                    thing = self.editable_tableWidget.item(i,j)
                    if thing is not None and thing.text() != '':
                        items = (thing.text())
                        x=str(items)
                        if j%5==1:
                            individual_dict['date']=x
                        elif j%5==2:
                            individual_dict['name']=x
                        elif j%5==3:
                            individual_dict['regno']=int(x)
                        elif j%5==4:
                            individual_dict['status']=x
                            counter = counter+1
                            whole[counter] = individual_dict
            print(whole)            
            for i in whole:
                getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
                print(whole[i])
                getdata.post_data(whole[i])

                        
    def upload_excelfile(self):
        self.empty_label1 = QLabel(self)
        self.empty_label1.setStyleSheet("color:black;background-color:white")
        self.empty_label1.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.upload_excel = QPushButton("Upload Excel",self)
        self.upload_excel.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.upload_excel.setFixedSize(120,30)
        self.attendance_label = QLabel("Students Attendance",self)
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label1.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.upload_excel)

        self.excel_attendance = QTableWidget(40,5,self)
        self.excel_attendance.setMaximumSize(1599,800)
        maxrow = 50
        maxcol = 5
        self.excel_attendance.setRowCount(maxrow) 
        self.excel_attendance.setColumnCount(maxcol)
        self.excel_attendance.verticalHeader().setVisible(False) 
        self.excel_attendance.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        def dialog():
            file = QFileDialog(self)
            self.dialog = file.getOpenFileName(self, 'Open attendance', '', '(*.xlsx)')
            print(self.dialog[1])
            path = self.dialog[0]
            wb_obj = openpyxl.load_workbook(path) 
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            sr = []
            regno = []
            name = []
            status = []
            counter = 2
            for i in range(1,sheet_obj.max_column):
                for j in range(3,sheet_obj.max_row-1):
                    cell_obj = sheet_obj.cell(row =j, column = i)
                    print(i,cell_obj.value)
                    if i == 1:
                        sr.append(cell_obj.value)
                    elif i == 2:
                        regno.append(cell_obj.value)
                    elif i == 3:
                        name.append(cell_obj.value)
                    elif i == 4:
                        status.append(cell_obj.value)
            #print(sr,regno,name,status)
            for i in range(0,len(sr)):
                for j in range(5):
                    item = QTableWidgetItem()
                    if j ==0:
                        item.setData(Qt.EditRole, sr[i])
                        self.excel_attendance.setItem(i,j, item)
                    elif j==1:
                        item.setData(Qt.EditRole,"12/24/2022")
                        self.excel_attendance.setItem(i,j, item)
                    elif j==2:
                        self.excel_attendance.resizeColumnToContents(i)
                        item.setData(Qt.EditRole,name[i])
                        self.excel_attendance.setItem(i,j, item)
                    
                    elif j==3:
                        item.setData(Qt.EditRole,regno[i])
                        self.excel_attendance.setItem(i,j, item)
                    elif j==4:
                        item.setData(Qt.EditRole,status[i])
                        self.excel_attendance.setItem(i,j, item)
                    
                    print(i,j,sr[i])

            whole = {}
            x=""
            counter = -1
            for i in range(maxrow):
                individual_dict = {}
                for j in range(maxcol):
                    print(i,j)
                    thing = self.excel_attendance.item(i,j)
                    if thing is not None and thing.text() != '':
                        items = (thing.text())
                        x=str(items)
                        if j%5==1:
                            individual_dict['date']=x
                        elif j%5==2:
                            individual_dict['name']=x
                        elif j%5==3:
                            individual_dict['regno']=int(x)
                        elif j%5==4:
                            individual_dict['status']=x
                            counter = counter+1
                            whole[counter] = individual_dict
            print(whole)
            self.send = QPushButton("SEND(TO SERVER)",self)
            self.send.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
            self.send.setFixedSize(120,30)
            self.hvbox.addWidget(self.send)
            self.upload_excel.hide()
            def set_database():                
                print("ok")
                #for i in whole:
                 #   getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
                  #  print(whole[i])
                   # getdata.post_data(whole[i])
            self.send.clicked.connect(lambda:set_database())


        self.upload_excel.clicked.connect(lambda:dialog())

        def QString(a):
            return (a)
        self.excel_attendance.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))
        self.gridlayout.addWidget(self.empty_label1,0,1,1,5)
        self.gridlayout.addWidget(self.excel_attendance,1,1,4,5)
    def showtable(self):
        self.empty_label = QLabel(self)
        self.empty_label.setStyleSheet("color:black;background-color:white")
        self.empty_label.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.edit_button = QPushButton("EDIT",self)
        self.edit_button.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.edit_button.setFixedSize(120,30)
        self.attendance_label = QLabel("Students Attendance",self)
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.edit_button)

        self.tableWidget = QTableWidget(40,5,self) 
        self.tableWidget.setMaximumSize(1500,800)        
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.verticalHeader().setVisible(False)

        getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
        result = getdata.get_response()
        print(result)
        ids,date,name,rollno,status= [],[],[],[],[]
        for i in range (len(result)):
            ids.append(result[i]['id'])
            date.append(result[i]['date'])
            name.append(result[i]['name'])
            rollno.append(result[i]['regno'])
            status.append(result[i]['status'])
        maxrow = 50
        maxcol = 5
        self.tableWidget.setRowCount(maxrow) 
        self.tableWidget.setColumnCount(maxcol)  
        currentRowCount = -1
        for i in range(len(ids)):
            currentRowCount += 1
            print("currentRowCount",currentRowCount)
            self.tableWidget.setItem(currentRowCount, 0, QTableWidgetItem(str(i+1)))
            self.tableWidget.setItem(currentRowCount, 1, QTableWidgetItem(str(date[i])))
            self.tableWidget.setItem(currentRowCount, 2, QTableWidgetItem(str(name[i])))
            self.tableWidget.setItem(currentRowCount, 3, QTableWidgetItem(str(rollno[i])))
            self.tableWidget.setItem(currentRowCount, 4, QTableWidgetItem(str(status[i])))

        def QString(a):
            return (a)
        self.tableWidget.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))
        self.gridlayout.addWidget(self.empty_label,0,1,1,5)
        self.gridlayout.addWidget(self.tableWidget,1,1,4,5)
    


    def showgradingcriteria(self):
        self.empty_label_for_grade = QLabel(self)
        self.empty_label_for_grade.setMaximumSize(1500,800)        
        self.empty_label_for_grade.setStyleSheet("background-color:white")
        self.gridlayout.addWidget(self.empty_label_for_grade,0,1,5,5)
        self.nestedgrid = QGridLayout(self)
        self.empty_label_for_grade.setLayout(self.nestedgrid)
        self.setcriteria = QLabel("𝐌𝐀𝐑𝐊𝐒 𝐃𝐢𝐬𝐭𝐫𝐢𝐛𝐮𝐭𝐢𝐨𝐧",self)
        self.setcriteria.setFont(QFont("arial",15))
        self.setcriteria.setAlignment(Qt.AlignCenter)

        self.vbox1 = QVBoxLayout(self)
        self.nestedlabel = QLabel(self)
        self.nestedlabel.setLayout(self.vbox1)
        self.quizlabel = QLabel("ADD QUIZ PERCENTAGE",self)
        self.lineedit_for_quiz = QLineEdit("0",self)
        self.lineedit_for_quiz.setFixedSize(150,30)      
        
        self.assignment = QLabel("ADD ASSIGNMENT PERCENTAGE",self)
        self.lineedit_for_assignment = QLineEdit("0",self)
        self.lineedit_for_assignment.setFixedSize(150,30)

        self.project = QLabel("ADD PROJECT PERCENTAGE",self)
        self.lineedit_for_project = QLineEdit("0",self)
        self.lineedit_for_project.setFixedSize(150,30)

        self.ohtlabel = QLabel("ADD OHT'S PERCENTAGE",self)
        self.lineedit_for_oht = QLineEdit("0",self)
        self.lineedit_for_oht.setFixedSize(150,30)

        self.final = QLabel("ADD Final PERCENTAGE",self)
        self.lineedit_for_final = QLineEdit("0",self)
        self.lineedit_for_final.setFixedSize(150,30)

        self.set = QPushButton("SET Marks Distribution",self)
        self.set.setStyleSheet('''QPushButton{background-color:lightgrey;color:black;border-radius:5px;border:2px solid grey;}
            QPushButton:hover{border:2px solid black;background-color:darkgrey}
            QPushButton:pressed{background-color:grey}''')
        self.set.setFixedSize(140,30)

        self.update = QLabel("TOTAL MUST BE = 100%",self)
        self.update.setFixedSize(150,160)
        self.update.setAlignment(Qt.AlignCenter)
        self.update.setFont(QFont("arial",10))


        self.vbox1.addStretch(1) 
        self.vbox1.addWidget(self.quizlabel)
        self.vbox1.addWidget(self.lineedit_for_quiz)
        self.vbox1.addWidget(self.assignment)
        self.vbox1.addWidget(self.lineedit_for_assignment)
        self.vbox1.addWidget(self.project)
        self.vbox1.addWidget(self.lineedit_for_project)
        self.vbox1.addWidget(self.ohtlabel)
        self.vbox1.addWidget(self.lineedit_for_oht)
        self.vbox1.addWidget(self.final)
        self.vbox1.addWidget(self.update)
        self.vbox1.addWidget(self.lineedit_for_final)
        self.vbox1.addWidget(self.update) 
        self.vbox1.addWidget(self.set)

        q =1
        a=1
        o=1
        f=1
        p=1
        series = QPieSeries()
        series.append("final", 1).setColor(QColor("skyblue"))
        series.append(" Assignments", 1).setColor(QColor("lightblue"))
        series.append("OHT's", 1).setColor(QColor("darkblue"))
        series.append("Project",1).setColor(QColor("green"))
        series.append("quiz", 1).setColor(QColor("lightgreen"))

        slice = QPieSlice()
        slice = series.slices()[0]
        slice.setExploded(True)
        slice.setLabelVisible(True)
        slice.setPen(QPen(Qt.red, 2))
        slice.setBrush(Qt.green)

 
        chart = QChart()
        chart.addSeries(series)
        chart.createDefaultAxes()
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setTitle("CHART") 
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)
        chartview = QChartView(chart)
        chartview.setRenderHint(QPainter.Antialiasing)
        #chartview.setMaximumSize(600,280)
        chart.createDefaultAxes()


        self.nestedgrid.addWidget(self.setcriteria,0,0,1,5)
        self.nestedgrid.addWidget(self.nestedlabel,1,0,4,1)
        self.nestedgrid.addWidget(chartview,1,1,1,1)

        def getpercentage():
            self.update.setWordWrap(True)
            answer = 0
            try:
                answer = int(self.lineedit_for_quiz.text()) + int(self.lineedit_for_assignment.text()) + int(self.lineedit_for_project.text()) + int(self.lineedit_for_oht.text()) + int(self.lineedit_for_final.text())
            except:
                self.update.setText("ENTER INT(PERCENTAGE VALUE)")
            if answer>100:
                self.update.setText("EXCEEDING 100%")
            elif answer<100:
                self.update.setText("BELOW 100%")
            else:
                print(answer)
                q = int(self.lineedit_for_quiz.text())
                a = int(self.lineedit_for_assignment.text())
                p = int(self.lineedit_for_project.text())
                o = int(self.lineedit_for_oht.text())
                f = int(self.lineedit_for_final.text())
                series.clear()
                series2 = QPieSeries()
                series2.append("Quiz", q).setColor(QColor("skyblue"))
                series2.append(" Assignments", a).setColor(QColor("lightblue"))
                series2.append("OHT's", o).setColor(QColor(QColor("darkblue")))
                series2.append("final", f).setColor(QColor("lightgreen"))
                series2.append("Project",p).setColor(QColor("green"))
                chart.addSeries(series2)

                slice = QPieSlice()
                input_list = []
                input_list.append(q)
                input_list.append(a)
                input_list.append(p)
                input_list.append(o)
                input_list.append(f)
                max_value = max(input_list)
                index = input_list.index(max_value)
                print(index)
                slice = series2.slices()[index]
                slice.setExploded(True)
                slice.setLabelVisible(True)
                slice.setPen(QPen(Qt.red, 2))
                slice.setBrush(Qt.green)
        self.set.clicked.connect(lambda:getpercentage())

    def FacultyMarks(self):
        self.grid = QGridLayout(self)
        self.empty_label = QLabel(self)
        self.empty_label.setStyleSheet("background-color:white")
        self.empty_label.setMaximumSize(1500,800)
        self.gridlayout.addWidget(self.empty_label,0,1,5,5)
        self.empty_label.setLayout(self.grid)

        self.label1 = QLabel("𝐄𝐯𝐚𝐥𝐮𝐚𝐭𝐢𝐨𝐧",self)
        self.label1.setFont(QFont("Serif",20))
        self.label1.setMaximumSize(1500,150)
        self.label1.setStyleSheet("color:white;background-color:skyblue")
        self.label1.setAlignment(Qt.AlignTop)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        self.label1.setGraphicsEffect(shadow)
        self.grid2 = QGridLayout(self)
        self.select_the_following = QLabel("",self)
        self.select_the_following.setLayout(self.grid2)
        shadow2 = QGraphicsDropShadowEffect()
        shadow2.setBlurRadius(15)
        self.select_the_following.setGraphicsEffect(shadow2)
        self.empty_forspace = QLabel(self)
        
        self.label3 = QLabel("select the following",self)
        self.label3.setStyleSheet("background-color:white")
        self.label3.setFont(QFont("Times",17))
        self.label3.setAlignment(Qt.AlignTop)


        self.label8 = QLabel("𝕯𝖊𝖕𝖆𝖗𝖙𝖒𝖊𝖓𝖙",self)
        self.label8.setFont(QFont("Times",11))

        self.label6 = QLabel("𝕭𝖆𝖙𝖈𝖍",self)
        self.label6.setFont(QFont("Times",11))


        self.label7 = QLabel("𝕾𝖊𝖈𝖙𝖎𝖔𝖓",self)
        self.label7.setFont(QFont("Times",11))
        
        self.ComboBox1=QComboBox(self)
        list1=["CS","EE"]
        self.ComboBox1.addItems(list1)
        self.ComboBox1.setMaximumSize(130,35)

        self.ComboBox2=QComboBox(self)
        self.ComboBox2.setMaximumSize(130,35)
        list_1=["CS-01","CS-02","CS-03"]
        list_2=["EE-01","EE-02","EE-03"]
        self.ComboBox2.addItems(list_1)
        self.sub_lists = [list_1, list_2]

        self.ComboBox3=QComboBox(self)
        list_11=["CS-01-A"]
        list_12=["CS-02-A","CS-02-B"]
        list_13=["CS-03-A"]
        self.ComboBox3.addItems(list_11)
        self.sub_lists1=[list_11,list_12,list_13]
        self.ComboBox3.setMaximumSize(130,35)

        self.label9 = QLabel("𝕾𝖚𝖇𝖏𝖊𝖈𝖙",self)
        self.label9.setFont(QFont("Times",11))

        self.ComboBox4=QComboBox(self)
        list_21=["Software Eng.","Software Eng Lab"]
        list_22=["DSA","DSA Lab"]
        list_23=["PF","PF Lab"]
        self.ComboBox4.addItems(list_21)
        self.sub_lists2=[list_21,list_22,list_23]
        self.ComboBox4.setMaximumSize(130,35)
        
        self.Search=QPushButton("𝐒𝐞𝐚𝐫𝐜𝐡",self)
        self.Search.setMaximumSize(150,30)
        self.Search.setStyleSheet('''QPushButton{background-color:white;border-radius:15%;border:1px solid black}
            QPushButton:hover{background-color:skyblue}
            QPushButton:pressed{padding-top:5px}''')
        self.Search.clicked.connect(self.Search_students)
        self.empty_space = QLabel(self)
        self.grid2.addWidget(self.label3,0,0,1,3)
        self.grid2.addWidget(self.label8,1,0,1,1)
        self.grid2.addWidget(self.label6,1,1,1,1)
        self.grid2.addWidget(self.label7,1,2,1,1)
        self.grid2.addWidget(self.ComboBox1,2,0,1,1)
        self.grid2.addWidget(self.ComboBox2,2,1,1,1)
        self.grid2.addWidget(self.ComboBox3,2,2,1,1)
        self.grid2.addWidget(self.label9,3,0,1,1)
        self.grid2.addWidget(self.ComboBox4,4,0,1,1)
        self.grid2.addWidget(self.empty_space,5,0,1,1)
        self.grid2.addWidget(self.Search,6,1,2,1)

        self.grid.addWidget(self.label1,0,0,2,20)
        self.grid.addWidget(self.select_the_following,1,1,5,18)
        self.grid.addWidget(self.empty_forspace,6,0,1,20)
        def activated(index):
            print("Activated index:", index)
        self.ComboBox1.activated.connect(activated)
        self.ComboBox2.activated.connect(activated)
        self.ComboBox3.activated.connect(activated)
        self.ComboBox4.activated.connect(activated)   
        def updateCombo(index):
            self.ComboBox2.clear()
            if str(self.ComboBox1.currentText()) == "CS":
                self.ComboBox2.addItems(self.sub_lists[0])
            else:
                self.ComboBox2.addItems(self.sub_lists[1])
        self.ComboBox1.currentIndexChanged.connect(updateCombo)
        

        def updateCombo1(index):
            self.ComboBox3.clear()
            if str(self.ComboBox2.currentText()) == "CS-01":
                self.ComboBox3.addItems(self.sub_lists1[0])
            elif str(self.ComboBox2.currentText()) == "CS-02":
                self.ComboBox3.addItems(self.sub_lists1[1])
            else:
                self.ComboBox3.addItems(self.sub_lists1[2])
        self.ComboBox2.currentIndexChanged.connect(updateCombo1)

        def updateCombo2(index):
            self.ComboBox4.clear()
            if str(self.ComboBox3.currentText()) == "CS-01-A":
                self.ComboBox4.addItems(self.sub_lists2[0])
            elif str(self.ComboBox3.currentText()) == "CS-03-A":
                self.ComboBox4.addItems(self.sub_lists2[2])
            elif str(self.ComboBox3.currentText()) == "CS-02-A"or "CS-02-B":
                self.ComboBox4.addItems(self.sub_lists2[1])
        self.ComboBox3.currentIndexChanged.connect(updateCombo2)
        
    def Search_students(self):
            if self.ComboBox1.currentText()=="CS" and self.ComboBox2.currentText()=="CS-02" and self.ComboBox3.currentText()=="CS-02-B" and self.ComboBox4.currentText()=="DSA":
                self.select_the_following.hide()
                self.select_student = QComboBox(self)
                self.select_student.setMaximumSize(200,35)
                students_list = ["Laraib Khalid","Sundas Fatima","Syed Muhammad Dawood","Fibha Ayaz","Laiba Qayoom","Amna Kanwal","Sufian Amin","Qazi Ahmed Abdullah","Maryam"
                ,"Abdullah Mudassir Anwer","Zoya Shafqat","Muhammad Haroon Siddique","Muhammad Saad Ali","Qinwan Azhar Aziz","Zainab Khalid","Muhammad Sameer Butt","Mujeeb Ahmad"
                ,"Tashfeen Basir","Mohammad Uzair","Baryal Khan Khattak","Muhammad Huzaifa Nasir","Laiba Ihsan","Mujtaba Shahid Malik","Abdullah Saqib","Attia Batool","Rehman Munawar"
                ,"Zinneerah Zawar Ghauri","Saad Ashraf","Muhammad Abdullah Ahmed Jamal","Maham Noor","Rao Ahmed Bin Aleem","Ali Jone","Malik Zeshan Ahmed","Tajamul Sarwar","Arzanish Kayani",
                "Muhammad Abdullah Abid","Ali Ahmad Bajwa","Nafey Javed Bukhari","Muhammad Adnan","Abdullah","Abdul Rahim"]
                self.select_student.addItems(students_list)
                self.select_student.setEditable(True)
                self.select_student.setInsertPolicy(QComboBox.NoInsert)
                self.select_student.completer().setCompletionMode(QCompleter.PopupCompletion)
                self.ComboBox5 = QComboBox(self)
                list_combo_5 = ["quiz","assignment","project","oht1","oht2","final"]
                self.ComboBox5.addItems(list_combo_5)
                self.ComboBox5.setMaximumSize(200,35)
                self.grid.addWidget(self.select_student,2,9,1,1)
                self.grid.addWidget(self.ComboBox5,3,9,1,1)
                
                def Search_func():
                    self.Search2 = QPushButton(self)
                    var = "ADD " + str(self.select_student.currentText()) +"'s"+" result" 
                    self.Search2.setText(var)
                    self.Search2.setStyleSheet('''QPushButton{background-color:white;border-radius:15%;border:1px solid black}
                QPushButton:hover{background-color:skyblue}
                QPushButton:pressed{padding-top:5px}''')
                    self.Search2.setMaximumSize(200,40)
                    self.grid.addWidget(self.Search2,4,9,2,1)
                    def check():
                        if self.ComboBox5.currentText()=="quiz":
                            self.label1.hide()
                            self.marks_entry()

                    self.Search2.clicked.connect(check)


                self.select_student.activated.connect(Search_func)

    def marks_entry(self):
        self.empty_label = QLabel(self)
        self.empty_label.setStyleSheet("color:black;background-color:white")
        self.empty_label.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.done = QPushButton("Done",self)
        self.done.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.done.setFixedSize(60,30)
        self.maxnum = QLabel("Max marks:",self)
        self.maxnum.setFixedSize(60,30)
        self.lineedit = QLineEdit(self)
        self.lineedit.setFixedSize(130,30)
        self.done.clicked.connect(lambda:afterdone())
     
        def afterdone():
            Faculty.date = self.lineedit2.text()
            Faculty.maxmarks = self.lineedit.text()
            self.lineedit.hide()
            self.lineedit2.hide()
            self.maxnum.hide()
            self.date.hide()
            self.done.hide()
            string = "Date : "+Faculty.date+"\t\t Maximum Marks of quiz :"+Faculty.maxmarks
            self.label_to_show_data = QLabel(self)
            self.label_to_show_data.setText(string)

            self.upload = QPushButton("upload",self)
            self.upload.setFixedSize(130,30)
            self.upload.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
            self.hvbox.addWidget(self.upload)
            self.hvbox.addWidget(self.label_to_show_data)
            self.upload.clicked.connect(lambda:uploadmarks())
            
        self.date = QLabel("Date:",self)
        self.date.setFixedSize(50,30)
        self.lineedit2 = QLineEdit(self)
        self.lineedit2.setFixedSize(130,30)

        self.attendance_label = QLabel("Students Marks Entry",self)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.date)
        self.hvbox.addWidget(self.lineedit2)
        self.hvbox.addWidget(self.maxnum)
        self.hvbox.addWidget(self.lineedit)
        self.hvbox.addWidget(self.done)


        self.tableWidget1 = QTableWidget(40,4,self)
        self.tableWidget1.setMaximumSize(1500,800)        
        self.tableWidget1.verticalHeader().setVisible(False) 
        self.tableWidget1.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)


        maxrow = 50
        maxcol = 4
        self.tableWidget1.setRowCount(maxrow) 
        self.tableWidget1.setColumnCount(maxcol)  


        def QString(a):
            return (a)
        self.tableWidget1.setHorizontalHeaderLabels(QString("Sr;Name;Rollno;Obtain Marks").split(";"))
        self.gridlayout.addWidget(self.empty_label,0,1,1,5)
        self.gridlayout.addWidget(self.tableWidget1,1,1,4,5)

        def uploadmarks():
            marks = {}
            x=""
            counter = -1
            for i in range(maxcol):
                individual_dict = {}
                for j in range(maxrow):

                    thing = self.tableWidget1.item(i,j)
                    if thing is not None and thing.text() != '':
                        items = (thing.text())
                        x=str(items)
                       # individual_dict['date'
                        print(j)
                        individual_dict['typeofexam'] = "Quiz"
                        individual_dict['date'] = str(Faculty.date)
                        if j%4==1:
                            individual_dict['name']=str(x)
                        elif j%4==2:
                            individual_dict['rollno']=int(x)
                        elif j%4==3:
                            individual_dict['maxnum']=int(Faculty.maxmarks)
                            individual_dict['obtainnum']=int(x)
                            counter = counter+1
                            marks[counter] = individual_dict

            print(marks)
            for i in marks:
                getdata = RequestAndResponse("http://127.0.0.1:8000/quiz marks/")
                print(marks[i])
                getdata.post_data(marks[i])


    def gradesMaker(self):
        self.grid3 = QHBoxLayout(self)
        self.empty_label2 = QLabel(self)
        self.empty_label2.setStyleSheet("background-color:lightgrey")
        self.empty_label2.setMaximumSize(1500,50)
        self.empty_label2.setLayout(self.grid3)
        self.SortAZ=QComboBox(self)
        self.SortAZ.setMaximumSize(180,30)
        self.SortAZ.setStyleSheet("background-color:white")        
        lst = ["Sort","find min,amx marks,Average","A-Z","Z-A","Asscending Rollno","Descending Rollno","Lowest to Highest Marks","Highest to Lowest Marks"]
        self.SortAZ.addItems(lst)
        self.SortAZ.currentIndexChanged.connect(self.max2)
        self.SortAZ.currentIndexChanged.connect(self.average)
        self.grid3.addWidget(self.SortAZ)

        self.label_to_show_data1 = QLabel("...",self)
        self.label_to_show_data1.setAlignment(Qt.AlignCenter)
        self.grid3.addWidget(self.label_to_show_data1)
        self.label_to_show_data1.setStyleSheet("border:2px solid white")

        self.label_to_show_data2 = QLabel("...",self)
        self.grid3.addWidget(self.label_to_show_data2)
        self.label_to_show_data2.setAlignment(Qt.AlignCenter)
        self.label_to_show_data2.setStyleSheet("border:2px solid white")

        self.label_to_show_data3 = QLabel("...",self)
        self.grid3.addWidget(self.label_to_show_data3)
        self.label_to_show_data3.setAlignment(Qt.AlignCenter)
        self.label_to_show_data3.setStyleSheet("border:2px solid white")


        self.generate_grades = QPushButton("Generate grade",self)
        self.generate_grades.setStyleSheet('''QPushButton{background-color:#C0C0C0;border-radius:25%;border:1px solid black;border-radius:10%}
            QPushButton:hover{background-color:white}
            QPushButton:pressed{padding-top:5px}''') 
        self.generate_grades.setMaximumSize(150,40)
        self.grid3.addWidget(self.generate_grades)
        self.generate_grades.clicked.connect(self.generateGrade)
        

        
        self.gradesA = QTableWidget(40,5,self)
        self.gradesA.setMaximumSize(1599,800)
        maxrow = 43
        maxcol =5
        self.gradesA.setRowCount(maxrow) 
        self.gradesA.setColumnCount(maxcol) 
        self.gradesA.verticalHeader().setVisible(False)
        self.gradesA.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        def QString1(b):
            return (b)
        self.gradesA.setHorizontalHeaderLabels(QString1("Name;Rollno;Total%;Obtained%;Grade;").split(";"))
        self.gridlayout.addWidget(self.empty_label2,0,1,1,5)
        self.gridlayout.addWidget(self.gradesA,1,1,4,5)
        self.loaddata()

    def loaddata(self,val=None):
        self.studs=["Laraib Khalid","Sundas Fatima","Syed Muhammad Dawood","Fibha Ayaz","Laiba Qayoom","Amna Kanwal","Sufian Amin","Qazi Ahmed Abdullah","Maryam"
                ,"Abdullah Mudassir Anwer","Zoya Shafqat","Muhammad Haroon Siddique","Muhammad Saad Ali","Qinwan Azhar Aziz","Zainab Khalid","Muhammad Sameer Butt","Mujeeb Ahmad"
                ,"Tashfeen Basir","Mohammad Uzair","Baryal Khan Khattak","Muhammad Huzaifa Nasir","Laiba Ihsan","Mujtaba Shahid Malik","Abdullah Saqib","Attia Batool","Rehman Munawar"
                ,"Zinneerah Zawar Ghauri","Saad Ashraf","Muhammad Abdullah Ahmed Jamal","Maham Noor","Rao Ahmed Bin Aleem","Ali Jone","Malik Zeshan Ahmed","Tajamul Sarwar","Arzanish Kayani",
                "Muhammad Abdullah Abid","Ali Ahmad Bajwa","Nafey Javed Bukhari","Muhammad Adnan","Abdullah","Abdul Rahim"]
        self.students_rollno =['210201001','210201002','210201004','210201006','210201008','210201010','210201012','210201014','210201016','210201018','210201020','210201022','210201024','210201028'
                ,'210201033','210201034','210201039','210201041','210201043','210201045','210201047','210201050','210201052','210201054','210201058','210201060','210201062','210201064','210201069','210201070','210201071','210201073'
                ,'210201075','210201077','210201081','210201083','210201087',"210201090","210201093","210201095","210201097"]
        total=["100",'100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100','100']
        self.obtained=[]
        gradesDo=[]

        
        for i in range(41):
            x = randint(30,95)
            self.obtained.append(x)
            print(self.obtained)
        currentRowCount1=-1
        for i in range(41):
            currentRowCount1 += 1
            print("currentRowCount",currentRowCount1)
            self.gradesA.setItem(currentRowCount1, 0, QTableWidgetItem(str(self.studs[i])))
            self.gradesA.setItem(currentRowCount1, 1, QTableWidgetItem(str(self.students_rollno[i])))
            self.gradesA.setItem(currentRowCount1, 2, QTableWidgetItem(str(total[i])))
            self.gradesA.setItem(currentRowCount1, 3, QTableWidgetItem(str(self.obtained[i])))
    

    def max2(self):
        if self.SortAZ.currentText() == "A-Z" or "Z-A" or "Asscending Rollno" or "Descending Rollno" or "Lowest to Highest Marks" or "Highest to Lowest Marks":
            n = len(self.studs)
            for i in range(n-1):
                for j in range(0,n-i-1):
                    if self.SortAZ.currentText() == "A-Z":
                        if self.studs[j] > self.studs[j + 1]:
                            self.studs[j], self.studs[j + 1] = self.studs[j + 1], self.studs[j]
                            self.students_rollno[j], self.students_rollno[j + 1] = self.students_rollno[j + 1], self.students_rollno[j]
                            self.obtained[j], self.obtained[j + 1] = self.obtained[j + 1], self.obtained[j]
                    elif self.SortAZ.currentText() == "Z-A":
                        if self.studs[j] < self.studs[j + 1]:
                            self.studs[j], self.studs[j + 1] = self.studs[j + 1], self.studs[j]
                            self.students_rollno[j], self.students_rollno[j + 1] = self.students_rollno[j + 1], self.students_rollno[j]
                            self.obtained[j], self.obtained[j + 1] = self.obtained[j + 1], self.obtained[j]
                    elif self.SortAZ.currentText() == "Descending Rollno":
                        if self.students_rollno[j] < self.students_rollno[j + 1]:
                            self.students_rollno[j], self.students_rollno[j + 1] = self.students_rollno[j + 1], self.students_rollno[j]
                            self.studs[j], self.studs[j + 1] = self.studs[j + 1], self.studs[j]
                            self.obtained[j], self.obtained[j + 1] = self.obtained[j + 1], self.obtained[j]
                    elif self.SortAZ.currentText() == "Asscending Rollno":
                        if self.students_rollno[j] > self.students_rollno[j + 1]:
                            self.students_rollno[j], self.students_rollno[j + 1] = self.students_rollno[j + 1], self.students_rollno[j]
                            self.studs[j], self.studs[j + 1] = self.studs[j + 1], self.studs[j]
                            self.obtained[j], self.obtained[j + 1] = self.obtained[j + 1], self.obtained[j]
                    elif self.SortAZ.currentText() == "Lowest to Highest Marks":
                        if self.obtained[j] > self.obtained[j + 1]:
                            self.obtained[j], self.obtained[j + 1] = self.obtained[j + 1], self.obtained[j]
                            self.studs[j], self.studs[j + 1] = self.studs[j + 1], self.studs[j]
                            self.students_rollno[j], self.students_rollno[j + 1] = self.students_rollno[j + 1], self.students_rollno[j]
                    elif self.SortAZ.currentText() == "Highest to Lowest Marks":
                        if self.obtained[j] < self.obtained[j + 1]:
                            self.obtained[j], self.obtained[j + 1] = self.obtained[j + 1], self.obtained[j]
                            self.studs[j], self.studs[j + 1] = self.studs[j + 1], self.studs[j]
                            self.students_rollno[j], self.students_rollno[j + 1] = self.students_rollno[j + 1], self.students_rollno[j]           
            currentRowCount1=-1
            for i in range(41):
                currentRowCount1 += 1
                print("currentRowCount",currentRowCount1)
                self.gradesA.setItem(currentRowCount1, 0, QTableWidgetItem(str(self.studs[i])))
                self.gradesA.setItem(currentRowCount1, 1, QTableWidgetItem(str(self.students_rollno[i])))
                self.gradesA.setItem(currentRowCount1, 3, QTableWidgetItem(str(self.obtained[i])))
        if self.SortAZ.currentText() == "find min,amx marks,Average":
            print("in")
            save_real_value = self.obtained
            def bubbleSort(arr):
                n = len(arr)
                for i in range(n-1):
                    for j in range(0,n-i-1):
                        if arr[j] > arr[j + 1]:
                            arr[j], arr[j + 1] = arr[j + 1], arr[j]
                return [arr[n-1],arr[0]]
            maxvalue = bubbleSort(save_real_value)
            self.label_to_show_data1.setText("\tMaximum% : "+str(maxvalue[0]))
            self.label_to_show_data2.setText("\tMinimum% : "+str(maxvalue[1]))
            #def Average(self,obtained):
    def average(self):
        x=round(sum(self.obtained) / len(self.obtained),2)
        self.label_to_show_data3.setText("\tAverage% : "+str(x))
        return x

    def generateGrade(self):
        self.SortAZ.setEnabled(False)
        gernade=[]
        for i in self.obtained:
            if i>=self.average() and i<=self.average()+4:
                gernade.append("B")
            elif i>=self.average()+4 and i<=self.average()+8:
                gernade.append("B+")
            elif i>=self.average()+8 and i<=self.average()+12:
                gernade.append("A-")
            elif i>=self.average()+12:
                gernade.append("A")
            elif i>=self.average()-4 and i<self.average():
                gernade.append("B-")
            elif i>=self.average()-8 and i<self.average()-4:
                gernade.append("C+")
            elif i>=self.average()-12  and i<self.average()-8:
                gernade.append("C")
            elif i-16>=self.average()-16 and i<self.average()-12:
                gernade.append("C-")
            elif i>=self.average()-20  and i<self.average()-16:
                gernade.append("D+")
            elif i>=self.average()-24  and i<self.average()-20:
                gernade.append("D-")
            else:
                gernade.append("F")
        currentRowCount1=-1
        for i in range(41):
            currentRowCount1 += 1
            self.gradesA.setItem(currentRowCount1, 0, QTableWidgetItem(str(self.studs[i])))
            self.gradesA.setItem(currentRowCount1, 1, QTableWidgetItem(str(self.students_rollno[i])))
            self.gradesA.setItem(currentRowCount1, 3, QTableWidgetItem(str(self.obtained[i])))
            self.gradesA.setItem(currentRowCount1, 4, QTableWidgetItem(str(gernade[i])))
class MplCanvas(FigureCanvasQTAgg):

    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot()
        self.axes.set_xlim([1, 35])        
        self.axes.set_ylim([0, 100])
        self.axes.set_xlabel("Credit Hours")
        self.axes.set_ylabel("percentage")

        
        super(MplCanvas, self).__init__(fig)
class Profile(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1000,640)
        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)

        self.background = QLabel(self)
        pixmap = QPixmap('profilebackground')
        self.background.setPixmap(pixmap)
        self.background.setScaledContents(True)
        self.background.setMaximumSize(1500,940)
        self.gridlayout.addWidget(self.background,0,0,6,3)
        self.gridlayout2 = QGridLayout(self)
        self.colorlabel = QLabel(self)
        self.colorlabel.setStyleSheet("background-color:#418484;border-radius:10%")
        self.colorlabel.setMaximumSize(2000,100)
        self.colorlabel.setLayout(self.gridlayout2)


        self.yourinfo = QLabel("Your info",self)
        self.yourinfo.setMaximumSize(350,50)
        self.yourinfo.setFont(QFont("arial",25))
        self.yourinfo.setStyleSheet("color:white")

        self.hbox2=QHBoxLayout(self)
        self.empty_for_passes = QLabel(self)
        self.empty_for_passes.setStyleSheet("background-color:#43BABA;border-radius:20%")
        self.empty_for_passes.setLayout(self.hbox2)
        
        self.changepass = QPushButton("change password",self)
        self.changepass.setIcon(QIcon("changepass"))
        self.changepass.setFixedSize(120,30)
        self.changepass.setStyleSheet('''QPushButton{background-color:transparent;color:white}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')
        def seepassword():
            self.empty_label2.hide()
            self.empty_label3.hide()
            self.sc.hide()

            self.vbox4 = QVBoxLayout(self)
            self.seepass_label = QLabel(self)
            self.seepass_label.setStyleSheet("background-color:#43BABA;border-radius:20%")
            self.seepass_label.setFixedSize(400,350)

            self.seepass_label.setLayout(self.vbox4)

            self.hbox3 = QHBoxLayout(self)
            self.hbox4 = QHBoxLayout(self)
            self.hbox5 = QHBoxLayout(self)
            self.nested_for_password = QLabel(self)
            self.nested_for_id = QLabel(self)
            self.nested_for_id.setLayout(self.hbox4)
            self.nested_for_type = QLabel(self)
            self.nested_for_type.setLayout(self.hbox5)
            self.nested_for_password.setLayout(self.hbox3)
            self.nestedseepass1 = QLabel("Password:",self)
            self.nestedseepass1.setFont(QFont("arial",15))
            self.nestedseepass2 = QLabel("210201073",self)
            self.nestedseepass2.setFont(QFont("arial",15))
            self.nestedseepassbtn = QPushButton("edit",self)
            self.nestedseepassbtn.setStyleSheet('''QPushButton{background-color:transparent;color:white}
            QPushButton:hover{color:grey;}
            QPushButton:pressed{padding-top:4px}''')
            self.nestedseepassbtn.setIcon(QIcon("edit"))
            def edit_pass():
                self.hbox6 = QHBoxLayout(self)
                self.nestedseepass2.setText("")
                self.nestedseepass2.setLayout(self.hbox6)
                self.linedit1 = QLineEdit(self)
                self.linedit1.setStyleSheet("background-color:white")
                self.ok1 = QPushButton("SET Pass",self)
                self.ok1.setStyleSheet('''QPushButton{background-color:transparent;color:white}
            QPushButton:hover{color:grey;}
            QPushButton:pressed{padding-top:4px}''')
                self.nestedseepassbtn.setEnabled(False)
                self.hbox6.addWidget(self.linedit1)
                self.hbox6.addWidget(self.ok1)
                def okpass():
                    self.nestedseepassbtn.setEnabled(True)
                self.ok1.clicked.connect(lambda:okpass())

                
            self.nestedseepassbtn.clicked.connect(lambda:edit_pass())
            self.nestedseeid1 = QLabel("Id:",self)
            self.nestedseeid1.setFont(QFont("arial",14))
            self.nestedseeid2 = QLabel("Id number",self)
            self.nestedseeid2.setFont(QFont("arial",14))
            self.nestedseeidbtn = QPushButton("edit2",self)
            self.nestedseeidbtn.setStyleSheet('''QPushButton{background-color:transparent;color:white}
            QPushButton:hover{color:grey;}
            QPushButton:pressed{padding-top:4px}''')
            self.nestedseeidbtn.setIcon(QIcon("edit"))
            self.nestedseetype1 = QLabel("Type:",self)
            self.nestedseetype1.setFont(QFont("arial",13))
            self.nestedseetype2 = QLabel("student",self)
            self.nestedseetype2.setFont(QFont("arial",13))
            self.nestedseetypebtn = QPushButton("edit3",self)
            self.nestedseetypebtn.setStyleSheet('''QPushButton{background-color:transparent;color:white}
            QPushButton:hover{color:grey;}
            QPushButton:pressed{padding-top:4px}''')
            self.nestedseetypebtn.setIcon(QIcon("edit"))
            self.hbox3.addWidget(self.nestedseepass1)
            self.hbox3.addWidget(self.nestedseepass2)
            self.hbox3.addWidget(self.nestedseepassbtn)
            self.hbox4.addWidget(self.nestedseeid1)
            self.hbox4.addWidget(self.nestedseeid2)
            self.hbox4.addWidget(self.nestedseeidbtn)
            self.hbox5.addWidget(self.nestedseetype1)
            self.hbox5.addWidget(self.nestedseetype2)
            self.hbox5.addWidget(self.nestedseetypebtn)

            
            self.vbox4.addWidget(self.nested_for_password)
            self.vbox4.addWidget(self.nested_for_id)
            self.vbox4.addWidget(self.nested_for_type)
            self.gridlayout.addWidget(self.seepass_label,2,1,3,1)
            
        self.changepass.clicked.connect(lambda:seepassword())
        self.seepass = QPushButton("see password",self)
        self.seepass.setIcon(QIcon("seepass"))
        self.seepass.setFixedSize(120,30)
        self.seepass.setStyleSheet('''QPushButton{background-color:transparent;color:white}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')
        


        

        self.manage_address = QPushButton("Home",self)
        self.manage_address.setIcon(QIcon("houselogo"))
        self.manage_address.setFixedSize(120,30)
        self.manage_address.setStyleSheet('''QPushButton{background-color:transparent;color:white}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')

        self.manage_address.clicked.connect(window_1.going_back_to_dashboard)
        

        self.hbox2.addWidget(self.manage_address)
        self.hbox2.addWidget(self.seepass)
        self.hbox2.addWidget(self.changepass)

        self.gridlayout2.addWidget(self.yourinfo,0,0)
        self.gridlayout2.addWidget(self.empty_for_passes,0,2)
       
        self.hbox1 = QHBoxLayout(self)
        self.empty_label1 = QLabel(self)
        self.empty_label1.setMaximumSize(300,40)
        self.empty_label1.setLayout(self.hbox1)
        self.profilebutton = QPushButton("Profile",self)
        self.profilebutton.setFixedSize(80,30)
        self.profilebutton.setStyleSheet('''QPushButton{background-color:transparent;}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')
        self.profilebutton.setFont(QFont("arial",10))
        self.profilebutton.clicked.connect(lambda:profileshow())
        self.sc = MplCanvas(self, width=5, height=4, dpi=90)

        def profileshow():
            self.empty_label2.show()
            self.empty_label3.hide()
            self.sc.hide()


        self.contactbutton = QPushButton("contact info",self)
        self.contactbutton.setFixedSize(80,30)
        self.contactbutton.setStyleSheet('''QPushButton{background-color:transparent;}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')
        self.contactbutton.setFont(QFont("arial",10))
        self.contactbutton.clicked.connect(lambda:contactinfo())
        
        self.empty_label3 = QLabel(self) 
        self.empty_label3.setMaximumSize(250,400)
        self.empty_label3.hide()       
        def contactinfo():
            self.empty_label3.show()
            self.empty_label2.hide()
            self.sc.hide()
            self.vbox3 = QVBoxLayout(self)
            self.empty_label3.setStyleSheet("text-decoration: underline;background-color:rgba(255,255,255, 0.1)")
            self.empty_label3.setLayout(self.vbox3)
            self.contactinfo = QLabel("𝗖𝗼𝗻𝘁𝗮𝗰𝘁 𝗜𝗻𝗳𝗼")
            self.contactinfo.setFont(QFont("arial",20))

            self.contactno = QLabel("Contact no : 0311-12414532")
            self.contactno.setFont(QFont("arial",10))

            self.facebook_id = QLabel("facebook id : uzairsalug")
            self.facebook_id.setFont(QFont("arial",10))

            self.insta_id = QLabel("instagram : nafaybhalu")
            self.insta_id.setFont(QFont("arial",10))

            self.home_no = QLabel("Home number : 03123123455")
            self.home_no.setFont(QFont("arial",10))

            self.vbox3.addWidget(self.contactinfo)
            self.vbox3.addWidget(self.contactno)
            self.vbox3.addWidget(self.home_no)
            self.vbox3.addWidget(self.facebook_id)
            self.vbox3.addWidget(self.insta_id)
            self.gridlayout.addWidget(self.empty_label3,2,1,2,1)
        
        self.acadamic_progress = QPushButton("Acadmic Progress",self)
        self.acadamic_progress.setFixedSize(120,30)
        self.acadamic_progress.setStyleSheet('''QPushButton{background-color:transparent;}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')
        self.acadamic_progress.setFont(QFont("arial",10))
        self.acadamic_progress.clicked.connect(lambda:acadamic_progress_show())        
        def acadamic_progress_show():
            getdata = RequestAndResponse("http://127.0.0.1:8000/quiz marks/")
            result = getdata.get_response()
            print(result)
            maximumnumber,obtainnumber = [],[]
            quizes = []
            for i in range (len(result)):
                if result[i]['typeofexam'] == "Quiz":
                    if result[i]['rollno'] == 210201073:
                        print("huoafdhoz")
                        maximumnumber.append(result[i]['maxnum'])
                        obtainnumber.append(result[i]['obtainnum'])
            for i in range (len(obtainnumber)):     
                print(maximumnumber,obtainnumber)
                calculations = obtainnumber[i]/maximumnumber[i]*100
                quizes.append(calculations)
            print(quizes)

            self.empty_label3.hide()
            self.empty_label2.hide()
            plt.xticks([1, 2, 3, 4, 5])
            plt.rcParams["figure.figsize"] = [7.00, 3.50]
            plt.rcParams["figure.autolayout"] = True
            plt.style.use('seaborn-v0_8')
            plt.rcParams.update({'figure.facecolor':'white'})
            plt.rcParams.update({'axes.facecolor':'lightblue'})
            self.sc = MplCanvas(self, width=5, height=4, dpi=90)
            dataanalysis = [[0,0,0,0,0], [0,0,0,0,0], [0,0,0,0,0], [0,0,0,0,0], [0,0,0,0,0]]
            counter = len(quizes)
            for i in range(1,len(dataanalysis)):
                try:
                    dataanalysis[i][1] = quizes[i-1]
                except:
                    print("caught")
            print(dataanalysis)
            df = pd.DataFrame(dataanalysis,columns=['Coal', 'DataStructure','Discrete','Professional Practice','Graph Theory'])
            df.plot(ax=self.sc.axes)
            #self.sc.setMaximumSize(2000,00)
            self.gridlayout.addWidget(self.sc,2,1,4,2)

        self.hbox1.addWidget(self.profilebutton)
        self.hbox1.addWidget(self.contactbutton) 
        self.hbox1.addWidget(self.acadamic_progress)
        
        self.vbox2 = QVBoxLayout(self)
        self.empty_for_photo = QLabel(self)
        self.empty_for_photo.setFixedSize(120,200)
        self.empty_for_photo.setLayout(self.vbox2)
        
        self.photo = QLabel(self)
        self.photo.setMaximumSize(100,100)
        self.photo.setStyleSheet('''QLabel{background-image:url(profileimg);border-radius:50%}''')


        self.edit_button = QPushButton("Edit image",self)
        self.edit_button.setFixedSize(100,30)
        self.edit_button.setStyleSheet('''QPushButton{background-color:transparent;}
        QPushButton:hover{color:grey;}
        QPushButton:pressed{padding-top:4px}''')
        #def edu():
            #for tl in QtWidgets.QApplication.topLevelWidgets():
             #   tl.hide()
        #elf.edit_button.clicked.connect(lambda:edu())

        self.vbox2.addWidget(self.photo)
        self.vbox2.addWidget(self.edit_button)
        self.empty_forspace = QLabel(self)

        self.name = QLabel("𝗔𝗹𝗶 𝗝𝗼𝗻𝗲",self)
        self.name.setFixedSize(200,50)
        self.name.setFont(QFont("arial",25))

        self.email = QLabel("E-mail : Alidada@gmail.com",self)
        self.email.setFont(QFont("arial",10))

        self.dob = QLabel("DOB : 3/12/1890",self)
        self.dob.setFont(QFont("arial",10))


        self.location = QLabel("Nationality : Pakistan",self)
        self.location.setFont(QFont("arial",10))

        self.language = QLabel("Language : Urdu",self)
        self.language.setFont(QFont("arial",10))

        self.degree = QLabel("Degree : Bachelor of Computer Sience",self)
        self.degree.setFont(QFont("arial",10))

        self.form = QVBoxLayout(self)
        self.empty_label2 = QLabel(self)
        self.empty_label2.setFixedSize(250,400)
        self.empty_label2.setStyleSheet("  text-decoration: underline;background-color:rgba(255,255,255, 0.1)")

        self.empty_label2.setLayout(self.form)
        self.form.addWidget(self.name)
        self.form.addWidget(self.dob)
        self.form.addWidget(self.email)
        self.form.addWidget(self.location)
        self.form.addWidget(self.language)
        self.form.addWidget(self.degree)

        self.empty_forspace2 = QLabel(self)
        self.empty_forspace2.setMaximumSize(100,200)

        self.gridlayout.addWidget(self.colorlabel,0,0,1,3)
        self.gridlayout.addWidget(self.empty_label1,1,0,1,2)
        self.gridlayout.addWidget(self.empty_for_photo,2,0,1,1)
        self.gridlayout.addWidget(self.empty_forspace2,4,0,2,1)
        self.gridlayout.addWidget(self.empty_label2,2,1,3,1)

class Feedback(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1000,640)
        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)        
        self.setStyleSheet( '''QScrollBar{ background-color: black}''')

        self.vbox = QVBoxLayout(self)
        self.empty1 = QLabel(self)
        self.empty1.setMaximumSize(800,600)
        self.empty1.setLayout(self.vbox)

        vbox2 = QVBoxLayout(self)
        vbox3 = QVBoxLayout(self)
        vbox4 = QVBoxLayout(self)
        self.feedbacktopics = QPushButton("total feedback topics ",self)
        self.feedbacktopics.setStyleSheet('''QPushButton{background-color:#606060;border-radius:10%;border:1px solid black}
            QPushButton:hover{background-color:#404040}
            QPushButton:pressed{padding-top:5px}''')
        self.feedbacktopics.setMaximumSize(500,200)
        self.feedbacktopics.setFont(QFont("arial",12))
        self.feedbacktopics.setLayout(vbox2)
        self.feedbacktopics.clicked.connect(self.feedbackquestions)
        self.feedbacktopics_label = QLabel("7",self)
        self.feedbacktopics_label.setAlignment(Qt.AlignCenter)
        self.feedbacktopics_label.setStyleSheet("border:2px solid transparent")
        vbox2.addStretch(1)
        vbox2.addWidget(self.feedbacktopics_label)
        self.feedbackanswered = QPushButton("feedback answered",self)
        self.feedbackanswered.setStyleSheet('''QPushButton{background-color:#A0A0A0;border-radius:10%;border:1px solid black}
            QPushButton:hover{background-color:#C0C0C0}
            QPushButton:pressed{padding-top:5px}''')
        self.feedbackanswered.setMaximumSize(500,200)
        self.feedbackanswered.setFont(QFont("arial",12))
        self.feedbackanswered.setLayout(vbox3)
        self.feedbackanswered_label = QLabel("0",self)
        self.feedbackanswered_label.setAlignment(Qt.AlignCenter)
        self.feedbackanswered_label.setStyleSheet("border:2px solid transparent")
        vbox3.addStretch(1)
        vbox3.addWidget(self.feedbackanswered_label)       
        self.feedbackunanswered = QPushButton("feedback unanswered",self)
        self.feedbackunanswered.setStyleSheet('''QPushButton{background-color:white;border-radius:10%;border:1px solid black}
            QPushButton:hover{background-color:#E0E0E0}
            QPushButton:pressed{padding-top:5px}''')
        self.feedbackunanswered.setMaximumSize(500,200)
        self.feedbackunanswered.setFont(QFont("arial",12))
        self.feedbackunanswered.setLayout(vbox4)
        self.feedbackunanswered_label = QLabel("7",self)
        self.feedbackunanswered_label.setAlignment(Qt.AlignCenter)
        self.feedbackunanswered_label.setStyleSheet("border:2px solid transparent")
        vbox4.addStretch(1)
        vbox4.addWidget(self.feedbackunanswered_label)       
        self.vbox.addWidget(self.feedbacktopics)
        self.vbox.addWidget(self.feedbackanswered)
        self.vbox.addWidget(self.feedbackunanswered)

        self.label = QLabel("None",self)
        pixmap = QPixmap('feedbackbackground')
        self.label.setPixmap(pixmap)
        self.label.setScaledContents(True)
        self.label.setMaximumSize(1500,840)
        self.label.setAlignment(Qt.AlignCenter)
        
        self.gridlayout.addWidget(self.label,0,0,8,6)
        
        self.photo = QLabel(self)
        self.photo.setStyleSheet("border:2px solid transparent")
        self.photo.setFixedSize(100,100)
        self.movie = QMovie("feedback.gif")
        self.photo.setMovie(self.movie)
        self.photo.setScaledContents(True)
        self.movie.setSpeed(200)
        self.movie.start()

        self.btn_back = QPushButton("<",self)
        self.btn_back.setFixedSize(50,50)
        self.btn_back.setStyleSheet('''QPushButton{background-color:#C0C0C0;border-radius:25%;border:1px solid black}
            QPushButton:hover{background-color:white}
            QPushButton:pressed{padding-right:5px}''')
        self.btn_back.clicked.connect(window_1.going_back_to_dashboard)



        self.options = QPushButton(self)
        self.options.setFixedSize(50,50)
        self.options.setStyleSheet('''QPushButton{background-image:url(threedots);border:2px solid white;border-radius:25%}
            QPushButton:hover{border:2px solid lightgrey}
            QPushButton:pressed{border:4px solid lightgrey}
            QPushButton::menu-indicator { image: none; }''')

        self.submit = QAction("submit",self)
        self.menu = QMenu()
        self.menu.setStyleSheet('''QMenu{background-color:darkgrey;}
            QMenu:item{color:black;}
            QMenu:item:selected{color:white;}
            QMenu:pressed{background-color:lightgrey}''')
        self.menu.addAction(self.submit)
        self.options.setMenu(self.menu)
        self.submit.triggered.connect(self.getfeedback)

        hbox1 = QHBoxLayout(self)
        self.whiteline = QLabel(self)
        self.whiteline.setStyleSheet("border:2px solid black")
        self.whiteline.setLayout(hbox1)
        
        hbox1.addWidget(self.btn_back)
        hbox1.addStretch(1)
        hbox1.addWidget(self.photo)
        hbox1.addStretch(1)
        hbox1.addWidget(self.options)

        self.gridlayout.addWidget(self.empty1,2,2,3,2)
        self.gridlayout.addWidget(self.whiteline,0,0,1,6)

    def feedbackquestions(self):
        self.blur_effect = QGraphicsBlurEffect()
        self.blur_effect.setBlurRadius(5)
        self.label.setGraphicsEffect(self.blur_effect)
        self.empty1.hide()
        self.questions_grid = QGridLayout(self)
        self.hbox_for_teachers= QHBoxLayout(self)
        self.empty2 = QLabel(self)
        self.empty2.setLayout(self.hbox_for_teachers)
        self.teacher_name = QLabel("Tufail Shah",self)
        self.teacher_name.setFont(QFont("arial",15))
        self.teacher_name.setAlignment(Qt.AlignVCenter)
        
        self.teacher_photo = QLabel(self)
        self.teacher_photo.setScaledContents(True)
        self.teacher_photo.setMaximumSize(50,50)
        self.teacher_photo.setStyleSheet("QLabel{background-image:url(sirtufail);border-radius:20%}")
        self.hbox_for_teachers.addWidget(self.teacher_photo)
        self.hbox_for_teachers.addWidget(self.teacher_name)
        self.questions_empty_label = QLabel(self)
        self.questions_empty_label.setFixedSize(1000,540)
        self.questions_empty_label.setFont(QFont("Didot",10))
        self.questions_empty_label.setLayout(self.questions_grid)
        self.quesions = QLabel("Give the following answer and be authentic.",self)
        self.quesions.setFont(QFont("Didot",13))
        self.quesions.setStyleSheet("border:1px solid lightgrey")

        self.scroll = QScrollArea(self)
        self.scroll.setWidget(self.questions_empty_label)
        self.scroll.setWidgetResizable(True)
        self.scroll.setStyleSheet("background-color:transparent")

        self.empty3 = QLabel(self)
        self.bad = QLabel("Disagree",self)
        self.normal = QLabel("Neutral",self)
        self.good = QLabel("Agree",self)
        self.verygood = QLabel("Strongly agree",self)
        

        self.q1 = QLabel("1.Course objectives were understandable for me.",self)
        self.q1.setFont(QFont("arial",10))
        self.q2 = QLabel("2.Lectures ,tests and assignments were effective to achieve course learning outcomes.",self)
        self.q2.setFont(QFont("arial",10))
        self.q3 = QLabel("3.Cheating was easy during exam ",self)
        self.q3.setFont(QFont("arial",10))
        self.q4 = QLabel("4.Instructor was strict ",self)
        self.q4.setFont(QFont("arial",10))
        self.q5 = QLabel("5.Instructor was regular ",self)
        self.q5.setFont(QFont("arial",10))
        self.q6 = QLabel("6.The instructor shows interest in helping student to learn",self)
        self.q6.setFont(QFont("arial",10))
        self.q7 = QLabel("7.The course was benifitial ",self)
        self.q7.setFont(QFont("arial",10))
        self.q8 = QLabel("8.The course gave me the confidence to work in subjective areas.",self)
        self.q8.setFont(QFont("arial",10))
        self.q9 = QLabel("9.The instructional material increased my knowlege and skills in the subjective manner.",self)
        self.q9.setFont(QFont("arial",10))
        self.q10 = QLabel("10.The instructor was well prepared to deliver the contents effectively .",self)
        self.q10.setFont(QFont("arial",10))
        bg1 = QButtonGroup(self)
        self.raioC1R1 = QRadioButton(self)
        self.raioC2R1 = QRadioButton(self)
        self.raioC3R1 = QRadioButton(self)
        self.raioC4R1 = QRadioButton(self)
        bg1.addButton(self.raioC1R1)
        bg1.addButton(self.raioC2R1)
        bg1.addButton(self.raioC3R1)
        bg1.addButton(self.raioC4R1)
        
        bg2 = QButtonGroup(self)
        self.raioC1R2 = QRadioButton(self)
        self.raioC2R2 = QRadioButton(self)
        self.raioC3R2 = QRadioButton(self)
        self.raioC4R2 = QRadioButton(self)
        bg2.addButton(self.raioC1R2)
        bg2.addButton(self.raioC2R2)
        bg2.addButton(self.raioC3R2)
        bg2.addButton(self.raioC4R2)

        bg3 = QButtonGroup(self)
        self.raioC1R3 = QRadioButton(self)
        self.raioC2R3 = QRadioButton(self)
        self.raioC3R3 = QRadioButton(self)
        self.raioC4R3 = QRadioButton(self)
        bg3.addButton(self.raioC1R3)
        bg3.addButton(self.raioC2R3)
        bg3.addButton(self.raioC3R3)
        bg3.addButton(self.raioC4R3)


        bg4 = QButtonGroup(self)
        self.raioC1R4 = QRadioButton(self)
        self.raioC2R4 = QRadioButton(self)
        self.raioC3R4 = QRadioButton(self)
        self.raioC4R4 = QRadioButton(self)
        bg4.addButton(self.raioC1R4)
        bg4.addButton(self.raioC2R4)
        bg4.addButton(self.raioC3R4)
        bg4.addButton(self.raioC4R4)

        bg5 = QButtonGroup(self)
        self.raioC1R5 = QRadioButton(self)
        self.raioC2R5 = QRadioButton(self)
        self.raioC3R5 = QRadioButton(self)
        self.raioC4R5 = QRadioButton(self)
        bg5.addButton(self.raioC1R5)
        bg5.addButton(self.raioC2R5)
        bg5.addButton(self.raioC3R5)
        bg5.addButton(self.raioC4R5)

        bg6 = QButtonGroup(self)
        self.raioC1R6 = QRadioButton(self)
        self.raioC2R6 = QRadioButton(self)
        self.raioC3R6 = QRadioButton(self)
        self.raioC4R6 = QRadioButton(self)
        bg6.addButton(self.raioC1R6)
        bg6.addButton(self.raioC2R6)
        bg6.addButton(self.raioC3R6)
        bg6.addButton(self.raioC4R6)

        bg7 = QButtonGroup(self)
        self.raioC1R7 = QRadioButton(self)
        self.raioC2R7 = QRadioButton(self)
        self.raioC3R7 = QRadioButton(self)
        self.raioC4R7 = QRadioButton(self)
        bg7.addButton(self.raioC1R7)
        bg7.addButton(self.raioC2R7)
        bg7.addButton(self.raioC3R7)
        bg7.addButton(self.raioC4R7)

        bg8 = QButtonGroup(self)
        self.raioC1R8 = QRadioButton(self)
        self.raioC2R8 = QRadioButton(self)
        self.raioC3R8 = QRadioButton(self)
        self.raioC4R8 = QRadioButton(self)
        bg8.addButton(self.raioC1R8)
        bg8.addButton(self.raioC2R8)
        bg8.addButton(self.raioC3R8)
        bg8.addButton(self.raioC4R8)

        bg9 = QButtonGroup(self)
        self.raioC1R9 = QRadioButton(self)
        self.raioC2R9 = QRadioButton(self)
        self.raioC3R9 = QRadioButton(self)
        self.raioC4R9 = QRadioButton(self)
        bg9.addButton(self.raioC1R9)
        bg9.addButton(self.raioC2R9)
        bg9.addButton(self.raioC3R9)
        bg9.addButton(self.raioC4R9)

        bg10 = QButtonGroup(self)
        self.raioC1R10 = QRadioButton(self)
        self.raioC2R10 = QRadioButton(self)
        self.raioC3R10 = QRadioButton(self)
        self.raioC4R10 = QRadioButton(self)
        bg10.addButton(self.raioC1R10)
        bg10.addButton(self.raioC2R10)
        bg10.addButton(self.raioC3R10)
        bg10.addButton(self.raioC4R10)
        
        
        
        self.questions_grid.addWidget(self.quesions,0,0,2,6)
        self.questions_grid.addWidget(self.empty3,0,1,2,2)
        self.questions_grid.addWidget(self.bad,0,2,2,1)
        self.questions_grid.addWidget(self.normal,0,3,2,1)
        self.questions_grid.addWidget(self.good,0,4,2,1)
        self.questions_grid.addWidget(self.verygood,0,5,2,1)
        self.questions_grid.addWidget(self.q1,2,0,2,1)
        self.questions_grid.addWidget(self.q2,3,0,2,1)
        self.questions_grid.addWidget(self.q3,4,0,2,1)
        self.questions_grid.addWidget(self.q4,5,0,2,1)
        self.questions_grid.addWidget(self.q5,6,0,2,1)
        self.questions_grid.addWidget(self.q6,7,0,2,1)
        self.questions_grid.addWidget(self.q7,8,0,2,1)
        self.questions_grid.addWidget(self.q8,9,0,2,1)
        self.questions_grid.addWidget(self.q9,10,0,2,1)
        self.questions_grid.addWidget(self.q10,11,0,2,1)

        self.questions_grid.addWidget(self.raioC1R1,2,2,2,1)
        self.questions_grid.addWidget(self.raioC2R1,2,3,2,1)
        self.questions_grid.addWidget(self.raioC3R1,2,4,2,1)
        self.questions_grid.addWidget(self.raioC4R1,2,5,2,1)

        self.questions_grid.addWidget(self.raioC1R2,3,2,2,1)
        self.questions_grid.addWidget(self.raioC2R2,3,3,2,1)
        self.questions_grid.addWidget(self.raioC3R2,3,4,2,1)
        self.questions_grid.addWidget(self.raioC4R2,3,5,2,1)

        self.questions_grid.addWidget(self.raioC1R3,4,2,2,1)
        self.questions_grid.addWidget(self.raioC2R3,4,3,2,1)
        self.questions_grid.addWidget(self.raioC3R3,4,4,2,1)
        self.questions_grid.addWidget(self.raioC4R3,4,5,2,1)

        self.questions_grid.addWidget(self.raioC1R4,5,2,2,1)
        self.questions_grid.addWidget(self.raioC2R4,5,3,2,1)
        self.questions_grid.addWidget(self.raioC3R4,5,4,2,1)
        self.questions_grid.addWidget(self.raioC4R4,5,5,2,1)

        self.questions_grid.addWidget(self.raioC1R5,6,2,2,1)
        self.questions_grid.addWidget(self.raioC2R5,6,3,2,1)
        self.questions_grid.addWidget(self.raioC3R5,6,4,2,1)
        self.questions_grid.addWidget(self.raioC4R5,6,5,2,1)

        self.questions_grid.addWidget(self.raioC1R6,7,2,2,1)
        self.questions_grid.addWidget(self.raioC2R6,7,3,2,1)
        self.questions_grid.addWidget(self.raioC3R6,7,4,2,1)
        self.questions_grid.addWidget(self.raioC4R6,7,5,2,1)

        self.questions_grid.addWidget(self.raioC1R7,8,2,2,1)
        self.questions_grid.addWidget(self.raioC2R7,8,3,2,1)
        self.questions_grid.addWidget(self.raioC3R7,8,4,2,1)
        self.questions_grid.addWidget(self.raioC4R7,8,5,2,1)

        self.questions_grid.addWidget(self.raioC1R8,9,2,2,1)
        self.questions_grid.addWidget(self.raioC2R8,9,3,2,1)
        self.questions_grid.addWidget(self.raioC3R8,9,4,2,1)
        self.questions_grid.addWidget(self.raioC4R8,9,5,2,1)

        self.questions_grid.addWidget(self.raioC1R9,10,2,2,1)
        self.questions_grid.addWidget(self.raioC2R9,10,3,2,1)
        self.questions_grid.addWidget(self.raioC3R9,10,4,2,1)
        self.questions_grid.addWidget(self.raioC4R9,10,5,2,1)

        self.questions_grid.addWidget(self.raioC1R10,11,2,2,1)
        self.questions_grid.addWidget(self.raioC2R10,11,3,2,1)
        self.questions_grid.addWidget(self.raioC3R10,11,4,2,1)
        self.questions_grid.addWidget(self.raioC4R10,11,5,2,1)

        self.gridlayout.addWidget(self.scroll,2,0,11,6)
        self.gridlayout.addWidget(self.empty2,1,0,1,6)

    def getfeedback(self):
        values = []
        if self.raioC1R1.isChecked() == True:
            values.append(0)
        if self.raioC2R1.isChecked() == True:
            values.append(1)
        if self.raioC3R1.isChecked() == True:
            values.append(2)
        if self.raioC4R1.isChecked() == True:
            values.append(3)
      
        if self.raioC1R2.isChecked() == True:
            values.append(0)
        if self.raioC2R2.isChecked() == True:
            values.append(1)
        if self.raioC3R2.isChecked() == True:
            values.append(2)
        if self.raioC4R2.isChecked() == True:
            values.append(3)

        if self.raioC1R3.isChecked() == True:
            values.append(3)
        if self.raioC2R3.isChecked() == True:
            values.append(2)
        if self.raioC3R3.isChecked() == True:
            values.append(0)
        if self.raioC4R3.isChecked() == True:
            values.append(0)
       
        if self.raioC1R4.isChecked() == True:
            values.append(1)
        if self.raioC2R4.isChecked() == True:
            values.append(1)
        if self.raioC3R4.isChecked() == True:
            values.append(2)
        if self.raioC4R4.isChecked() == True:
            values.append(2)

        if self.raioC1R5.isChecked() == True:
            values.append(0)
        if self.raioC2R5.isChecked() == True:
            values.append(1)
        if self.raioC3R5.isChecked() == True:
            values.append(2)
        if self.raioC4R5.isChecked() == True:
            values.append(3)


        if self.raioC1R6.isChecked() == True:
            values.append(0)
        if self.raioC2R6.isChecked() == True:
            values.append(1)
        if self.raioC3R6.isChecked() == True:
            values.append(2)
        if self.raioC4R6.isChecked() == True:
            values.append(3)

        if self.raioC1R7.isChecked() == True:
            values.append(0)
        if self.raioC2R7.isChecked() == True:
            values.append(1)
        if self.raioC3R7.isChecked() == True:
            values.append(2)
        if self.raioC4R7.isChecked() == True:
            values.append(3)
        
        if self.raioC1R8.isChecked() == True:
            values.append(0)
        if self.raioC2R8.isChecked() == True:
            values.append(1)
        if self.raioC3R8.isChecked() == True:
            values.append(2)
        if self.raioC4R8.isChecked() == True:
            values.append(3)

        if self.raioC1R9.isChecked() == True:
            values.append(0)
        if self.raioC2R9.isChecked() == True:
            values.append(1)
        if self.raioC3R9.isChecked() == True:
            values.append(2)
        if self.raioC4R9.isChecked() == True:
            values.append(3)
       
        if self.raioC1R10.isChecked() == True:
            values.append(0)
        if self.raioC2R10.isChecked() == True:
            values.append(1)
        if self.raioC3R10.isChecked() == True:
            values.append(2)
        if self.raioC4R10.isChecked() == True:
            values.append(3)
        self.feedbackanswered_label.setText("1")
        self.feedbackunanswered_label.setText("6")
        self.scroll.hide()
        self.empty1.show()
        score = sum(values)
        print(score)
        return score

class Department(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1000,640)
        self.Widget = QWidget(self)
        self.setCentralWidget(self.Widget)
        self.layout = QGridLayout(self.Widget)
        self.firstscreen()
    def firstscreen(self):
        self.empty0 = QLabel(self)
        self.hv0 = QHBoxLayout(self)
        self.empty0.setLayout(self.hv0)
        self.head_label = QLabel("Department ",self)
        self.head_label.setFont(QFont("Times", 13,))
        self.head_label.setAlignment(Qt.AlignCenter)
        self.empty0.setStyleSheet("color: black; background-color:skyblue;border-radius:10px;border:2px; black")
        self.empty0.setMaximumSize(1910,80)
        self.back_btn = QPushButton("<",self)
        self.back_btn.setFixedSize(40,40)
        self.back_btn.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-left:5px}''')
        self.back_btn.clicked.connect(window_1.going_back_to_dashboard)
        self.hv0.addWidget(self.back_btn)
        self.hv0.addStretch(4)
        self.hv0.addWidget(self.head_label)
        self.hv0.addStretch(5)

        self.label = QLabel(self)
        self.label.setStyleSheet("background-color:transparent;")
        self.label.setPixmap(QPixmap("csdep.png"))
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setScaledContents(True)
        self.label.setMaximumSize(300,300)

        self.istlabel = QLabel(self)
        self.istlabel.setStyleSheet("background-color:transparent;")
        self.istlabel.setPixmap(QPixmap("ist-img.png"))
        self.istlabel.setAlignment(Qt.AlignRight)
        self.istlabel.setScaledContents(True)
        self.istlabel.setMaximumSize(250,150)

        self.flaglabel = QLabel(self)
        self.flaglabel.setStyleSheet("background-color:transparent")
        self.flaglabel.setPixmap(QPixmap("flag.png"))
        self.flaglabel.setScaledContents(True)
        self.flaglabel.setMaximumSize(250,200)

        

        self.bottom_label = QLabel(" ",self)
        self.bottom_label.setFont(QFont("Times", 13,))
        self.bottom_label.setStyleSheet("color: white; background-color:skyblue;border-radius:10px;border:2px; solid black")
        self.bottom_label.setMaximumSize(1910,80)


        self.button = QPushButton(" Notice board ",self)
        self.button.move(50,100)
        self.button.resize(250,100)
        self.button.setFont(QFont("Times",10))

        self.button.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button.setMaximumSize(250,100)
    
        self.button1 = QPushButton(" Faculty ",self)
        self.button1.setFont(QFont("Times",10))
        self.button1.setStyleSheet("")
        self.button1.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button1.setMaximumSize(250,100)

        self.button2 = QPushButton(" Cirriculum ",self)
        self.button2.setFont(QFont("Times",10))
        self.button2.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button2.setMaximumSize(250,100)

        self.button3 = QPushButton(" Check Whats new! ",self)
        self.button3.setFont(QFont("Times",10))
        self.button3.setStyleSheet("")
        self.button3.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button3.setMaximumSize(850,100)

        self.button7 = QPushButton(" View sporting activities! ",self)
        self.button7.setFont(QFont("Times",10))
        self.button7.setStyleSheet("")
        self.button7.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button7.setMaximumSize(850,100)

        self.button4 = QPushButton(" Projects ",self)
        self.button4.setFont(QFont("Times",10))
        self.button4.setStyleSheet("")
        self.button4.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button4.setMaximumSize(250,100)
        self.button4.clicked.connect(self.project)

        self.button5 = QPushButton(" Internships ",self)
        self.button5.setFont(QFont("Times",10))
        self.button5.setStyleSheet("")
        self.button5.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button5.setMaximumSize(250,100)

        

        self.button6 = QPushButton(" Societies ",self)
        self.button6.setFont(QFont("Times",10))
        self.button6.setStyleSheet("")
        self.button6.setStyleSheet('''QPushButton{background-color: skyblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{background-color: transparent;padding-top:5px}''')
        self.button6.setMaximumSize(250,100)



        self.hv = QHBoxLayout(self)
        self.empty = QLabel(self)
        self.empty.setLayout(self.hv)
        self.hv.addWidget(self.button)
        self.hv.addWidget(self.button1)
        self.hv.addWidget(self.button2)
        self.hv.addWidget(self.button4)
        self.hv.addWidget(self.button5)
        self.hv.addWidget(self.button6)

        self.hv1 = QHBoxLayout(self)
        self.empty1 = QLabel(self)
        self.empty1.setLayout(self.hv1)
        self.hv1.addWidget(self.istlabel)
        self.hv1.addWidget(self.label)
        self.hv1.addWidget(self.flaglabel)

        self.hv2 = QHBoxLayout(self)
        self.empty2 = QLabel(self)
        self.empty2.setLayout(self.hv2)
        self.hv2.addWidget(self.button3)
        self.hv2.addWidget(self.button7)


        self.layout.addWidget(self.empty0,0,0,1,6)
        self.layout.addWidget(self.empty,1,0,1,6)
        self.layout.addWidget(self.empty1,3,0,4,6)
        self.layout.addWidget(self.empty2,7,0,1,6)
        self.layout.addWidget(self.bottom_label,8,0,1,6)

    def project(self):
        self.Widget = QWidget(self)
        self.setCentralWidget(self.Widget)
        self.layout = QGridLayout(self.Widget)

        self.setStyleSheet("background-color:white")
        
        self.label = QLabel("𝐀𝐯𝐚𝐢𝐥𝐚𝐛𝐥𝐞 𝐏𝐫𝐨𝐣𝐞𝐜𝐭𝐬",self)
        self.label.setFont(QFont("arial", 13,))
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("color: black; background-color:transparent")
        self.label.setFont(QFont("arial",25))

        self.back =QPushButton("<",self)
        self.back.setMaximumSize(100,50)
        self.back.setStyleSheet('''QPushButton {
    background-color: lightblue;
    color: #FFFFFF;
    border-style: outset;
    padding: 2px;
    font: bold 20px;
    border-width: 6px;
    border-radius: 10px;
    border-color: ;
}
QPushButton:hover {
    background-color: skyblue;color:black;;
}

}
QPushButton:pressed{padding-left:6px}''')
        self.back.clicked.connect(self.firstscreen)

        self.grid2 = QGridLayout(self)
        self.empty = QLabel(self) 

        self.empty.setLayout(self.grid2)

        self.label1 = QLabel("Build a dashboard for IST \n Make traffic management application \n Reconfigue IT Lab Pc's \n Lead a flutter course for CS-04 ",self)
        self.label1.setFont(QFont("arial", 13,))
        self.label1.setAlignment(Qt.AlignCenter)
        self.label1.setStyleSheet('''QLabel{color: black; background-color:lightblue;border-radius:10px;border:15px solid white;border-style: double}''')


        self.button = QPushButton("Submit a project",self)
        self.button.setFont(QFont("arial",13))
        self.button.setStyleSheet('''QPushButton{background-color: lightblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding-top:4px;}''')
        self.button.setMaximumSize(350,40)


        self.button1 = QPushButton("Request a project",self)
        self.button1.setFont(QFont("arial",13))
        self.button1.setStyleSheet('''QPushButton{background-color: lightblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding-top:4px;}''')
        self.button1.setMaximumSize(350,40)

        self.button2 = QPushButton("Create a project",self)
        self.button2.setFont(QFont("arial",13))
        self.button2.setStyleSheet('''QPushButton{background-color: lightblue;border-radius:10px;border:2px solid black;}
        QPushButton:hover{background-color:white;}
        QPushButton:pressed{padding-top:4px;}''')
        self.button2.setMaximumSize(350,40)

        self.ver = QVBoxLayout(self)
        self.empty2 = QLabel(self)
        self.empty2.setStyleSheet("border :10px solid lightblue;border-style:double")
        self.empty2.setLayout(self.ver)
        self.ver.addWidget(self.button)
        self.ver.addWidget(self.button1)
        self.ver.addWidget(self.button2)


        self.grid2.addWidget(self.label1,0,0,5,5)
        self.grid2.addWidget(self.label,0,2)
        self.grid2.addWidget(self.back,0,0)
        self.grid2.addWidget(self.empty2,1,5,3,2)
        
        self.layout.addWidget(self.empty,0,0)
        #self.layout.addWidget()

class Events(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1000,640)
        self.Widget = QWidget(self)
        self.setCentralWidget(self.Widget)
        self.layout = QGridLayout(self.Widget)
        self.setStyleSheet("QScrollBar{ background-color: transparent}QMainWindow{background-color:#a7573a}")

        self.background = QLabel(self)
        pixmap = QPixmap('evback')
        self.background.setPixmap(pixmap)
        self.background.setScaledContents(True)
        
        self.back_btn = QPushButton("",self)
        self.back_btn.setFixedSize(130,70)
        self.back_btn.setStyleSheet('''QPushButton{background-color: transparent;border-radius:35%;border:2px solid transparent;}''')  
        self.back_btn.clicked.connect(window_1.going_back_to_dashboard)
        self.hv1 = QGridLayout(self)
        self.empty3 = QLabel(self)  
        self.empty3.setMaximumSize(150,100) 
        self.empty3.setStyleSheet('''QLabel{background-color: transparent;border-radius:35%;border:2px solid transparent;}
        QLabel:hover{background-color :#a7573a;}''')    
        self.empty3.setLayout(self.hv1)
        self.movie2 = QMovie("backbtn.gif")
        self.empty3.setMovie(self.movie2)
        self.hv1.addWidget(self.back_btn)
        self.empty3.setScaledContents(True)
        self.movie2.setSpeed(200)
        self.movie2.start()

        self.calendar = QCalendarWidget(self)
        self.calendar.setMaximumSize(400,150)
        self.calendar.setStyleSheet('''QMenu {
    background: #a7573a;
}
QMenu::item:selected {
    background: yellow;
    border-radius: 2px;
}
#qt_calendar_navigationbar {
    background: #a7573a
}
QCalendar{background-color:#a7573a}
QCalendarWidget QTableView
{
    background-color:white;
}''')
        self.hv = QHBoxLayout(self)
        self.label1 = QLabel(self)
        self.label1.setLayout(self.hv)
        pixmap2 = QPixmap('even.png')
        self.label1.setPixmap(pixmap2)
        self.label1.setScaledContents(True)
        self.label1.setMaximumSize(360,150)


        string = ''' ->Freelancing by Nafay Jawed | 2pm | (10/01/23) | Auditorium\n-> Raashan Drive | 9am | (14/01/23) | Auditorium\n-> Qawali Night | 5pm | (15/01/23) | Raza block\n-> Signal Traffic |11am| (20/01/23) | Block 6'''

        self.empty1 = QLabel(self)
        self.empty1.setText(string)
        self.empty1.setStyleSheet("background-color:#fcf7a3;border:20px transparent")
        self.empty1.setFont(QFont("arial", 15,))
        self.empty1.setFixedSize(1500,650)
        self.empty1.setAlignment(Qt.AlignTop)
        self.empty1.setWordWrap(True)
        self.empty = QScrollArea(self)
        self.empty.setMaximumSize(1500,450)
        self.empty.setWidget(self.empty1)
        
        self.label = QLabel("None",self)
        self.label.setStyleSheet("background-color:white;border-radius:30%;border:10px solid white")
        self.label.setMaximumSize(380,80)
        self.label.setAlignment(Qt.AlignCenter)
        self.movie = QMovie("rain.gif")
        self.label.setMovie(self.movie)
        self.label.setScaledContents(True)
        self.movie.setSpeed(200)
        self.movie.start()

        self.lcdNumber = QLabel("up next:",self)
        self.lcdNumber.setStyleSheet("color: red;background-color:white;")
        self.lcdNumber.setMaximumSize(200,60)
        self.lcdNumber.setFont(QFont("arial",10))
        self.lcdNumber.setWordWrap(True)
        self.lcdNumber_seconds = QLabel("seconds",self)
        self.lcdNumber_seconds.setStyleSheet("color: black;background-color:white;padding-bottom:5px")
        self.lcdNumber_seconds.setMaximumSize(200,30)
        self.lcdNumber_seconds.setFont(QFont("ariel",24))
        self.hv2 = QVBoxLayout(self)
        self.empty4 = QLabel(self)
        self.empty4.setLayout(self.hv2)
        self.hv2.addWidget(self.lcdNumber_seconds)
        self.hv2.addWidget(self.lcdNumber)

        self.lcdNumber_seconds2 = QLabel(self)
        self.lcdNumber_seconds2.setStyleSheet("color: black;background-color:white;padding-bottom:5px")
        self.lcdNumber_seconds2.setMaximumSize(200,30)
        self.lcdNumber_seconds2.setFont(QFont("ariel",24))

        self.remaining = QLabel("Remaining Time:",self)
        self.remaining.setStyleSheet("color: black;background-color:white;padding-bottom:5px")
        self.remaining.setMaximumSize(200,30)
        self.remaining.setFont(QFont("ariel",12))
        
        self.hv3 = QVBoxLayout(self)
        self.empty5 = QLabel(self)
        self.empty5.setLayout(self.hv3)
        self.hv3.addWidget(self.remaining)
        self.hv3.addWidget(self.lcdNumber_seconds2)

        self.layout.addWidget(self.background,0,0,5,9)
        self.layout.addWidget(self.empty3,0,0,1,1)
        self.layout.addWidget(self.empty4,2,0,1,1)
        self.layout.addWidget(self.calendar,0,7,1,2)
        self.layout.addWidget(self.label,0,3,1,4)
        self.layout.addWidget(self.label1,0,3,1,6)
        self.layout.addWidget(self.empty,2,1,3,7)   
        self.layout.addWidget(self.empty5,2,8,1,1)  


        queue = ["Freelancing by Nafay Javed |2|pm ","Raashan Drive |9|am ","> Qawali Night |5|pm ","Signal Traffic |11|am"]
        self.lcdNumber.setText("next: "+queue[0])
        y = ""
        for i in queue:
            for j in i:
                x = i.split("|")
                if x[1] not in y:
                    y = y+","+x[1]
        
        timequeue = y.split(",")
        timequeue.pop(0)
        print(timequeue)

        def showtime():
            current_time = QTime.currentTime()
            label_time = current_time.toString('hh:mm:ss')
            self.lcdNumber_seconds.setText(label_time)
            print("hour:",current_time.hour())
            print("minute:",current_time.minute())

        timer = QTimer(self)
        timer.timeout.connect(lambda:showtime())  
        timer.start(1000)

        current_time = QTime.currentTime()
        label_time = current_time.hour() - int(timequeue[0])
        label = label_time * 3600
        self.countdown=label
        def showremaining2():    

            if self.countdown == 0:
                timer2.stop()
                timequeue.pop(0)
                queue.pop(0)
            print("hi")
            self.countdown = self.countdown - 1
            self.lcdNumber_seconds2.setText(str(self.countdown)+"'s")

        timer2 = QTimer(self)
        timer2.timeout.connect(lambda:showremaining2())  
        timer2.start(1000)

class Messages(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1000,640)
        self.Widget = QWidget(self)
        self.setCentralWidget(self.Widget)
        self.layout = QGridLayout(self.Widget)

        self.btn = QPushButton(self)
        self.btn.setStyleSheet(('''QPushButton{background-image:url(backmsg);border-radius:25%}QPushButton:pressed{border:1px solid white}'''))
        self.btn.setMaximumSize(90,90)
        self.btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn.clicked.connect(window_1.going_back_to_dashboard)

        self.background = QLabel("None",self)
        self.background.setStyleSheet("background-color:black;")
        self.background.setMaximumSize(2000,840)
        self.background.setAlignment(Qt.AlignCenter)

        self.topic = QLabel(self)       
        pixmap = QPixmap('istmessages')
        self.topic.setPixmap(pixmap)
        self.topic.setScaledContents(True)
        self.topic.setMaximumSize(350,200)

        self.movie = QMovie("loader2.gif")
        self.background.setMovie(self.movie)
        self.background.setScaledContents(True)
        self.movie.setSpeed(200)
        self.movie.start()

        self.hv = QHBoxLayout(self)
        self.empty1 = QLabel(self)
        self.empty1.setLayout(self.hv)
        self.empty1.setStyleSheet("border:2px solid transparent")
        self.label = QLabel("Leave Message:")
        self.messages_pic = QLabel(self)        
        pixmap = QPixmap('formsg')
        self.messages_pic.setPixmap(pixmap)
        self.messages_pic.setScaledContents(True)
        self.messages_pic.setMaximumSize(100,100)
        self.hv.addWidget(self.label)
        self.hv.addWidget(self.messages_pic)


        self.vb = QVBoxLayout(self)
        self.empty0 = QLabel(self)
        self.empty0.setMaximumSize(900,350)
        self.empty0.setStyleSheet("background-color: rgba(255, 255, 255, 90);border: 2px solid black")
        self.empty0.setLayout(self.vb)
        self.textEdit1 = QTextEdit("Write",self)
        self.textEdit1.setMaximumSize(600,150)
        self.textEdit1.setStyleSheet("border:2px solid black")

        self.hv1 = QHBoxLayout(self)
        self.empty2 = QLabel(self)
        self.empty2.setLayout(self.hv1)
        self.empty2.setStyleSheet("border:2px solid transparent")
        self.send = QPushButton("SEND",self)
        self.send.setMaximumSize(100,30)
        self.send.setStyleSheet('''QPushButton{background-color: rgba(255, 255, 255, 90);border: 2px solid lightblue}QPushButton:hover{background-color:rgba(220,255,255);}
        QPushButton:pressed{padding-top:2px;background-color:white}''')
        self.send.clicked.connect(self.sending_message)
        self.clear = QPushButton("Clear",self)
        self.clear.setMaximumSize(100,30)
        self.clear.setStyleSheet('''QPushButton{background-color: rgba(255, 255, 255, 90);border: 2px solid lightblue}QPushButton:hover{background-color:rgba(220,255,255)}
        QPushButton:pressed{padding-top:2px;background-color:white}''')
        self.clear.clicked.connect(self.deletemsg)
        self.hv1.addWidget(self.send)
        self.hv1.addWidget(self.clear)
        self.vb.addWidget(self.empty1)
        self.vb.addWidget(self.textEdit1)
        self.vb.addWidget(self.empty2)

        self.empty1 = QLabel(self)
        self.empty1.setText("string")
        self.empty1.setStyleSheet('''QLabel{background-color: rgba(255, 255, 255, 90);border: 20px solid lightgrey}''')
        self.empty1.setFont(QFont("arial", 15,))
        self.empty1.setFixedSize(1500,650)
        self.empty1.setAlignment(Qt.AlignTop)
        self.empty1.setWordWrap(True)
        self.empty = QScrollArea(self)
        self.empty.setMaximumSize(900,100)
        self.empty.setWidget(self.empty1)

        self.layout.addWidget(self.background,0,0,6,6)
        self.layout.addWidget(self.btn,0,0,1,2)
        self.layout.addWidget(self.topic,0,2,1,1)
        self.layout.addWidget(self.empty,1,1,1,4)
        self.layout.addWidget(self.empty0,3,1,1,4)
        self.data = RequestAndResponse("http://127.0.0.1:8000/ chats betweem student/teacher/")
        self.get_msgs()
    def get_msgs(self):
        answer = self.data.get_response()
        print(answer)
        chat = ""
        name = ""
        final = ""
        for i in range(len(answer)):
            if answer[i]["name"]:
                name = answer[i]["name"]
            if answer[i]["chat"]:
                chat = answer[i]["chat"]
            final = final +  name +": "+ chat+"\n"
            self.empty1.setText(final)
    def sending_message(self):

        if window_1.returning()[0] == "210201073":
            result = self.textEdit1.toPlainText()
            print(result)
            dict1={"name":"ALI JONE","chat":result}
            print(dict1)
            self.data.post_data(dict1)
            self.get_msgs()
        elif window_1.returning()[0] != "210201073":
            result = self.textEdit1.toPlainText()
            print(result)
            dict1={"name":"Sir Nadeem","chat":result}
            print(dict1)
            self.data.post_data(dict1)
    def deletemsg(self):
        if window_1.returning()[0] == "210201073":
            self.data2 = RequestAndResponse("http://127.0.0.1:8000/ chats betweem student/teacher/")
            self.data2.delete_user("ALI JONE")
            answer = self.data.get_response()
            print(answer)
            if answer == [] or None :
                self.empty1.setText("No chat available")    
        elif window_1.returning()[0] != "210201073":

            self.data2 = RequestAndResponse("http://127.0.0.1:8000/ chats betweem student/teacher/")
            self.data2.delete_user("Sir Nadeem")
            answer = self.data.get_response()
            print(answer)
            if answer == [] or None :
                self.empty1.setText("No chat available")

def splash():
    app = QApplication(sys.argv)
    splash = QSplashScreen(QPixmap('logo_ist'))
    splash.show()    
    return app.exec_()

      
desktop_app = QApplication(sys.argv)                  ## sys.arvg is a list in python which contains command lines arguements    
window_1 = Ist_Ams()
stack = QStackedWidget()
stack.addWidget(window_1)

window_1.Gui_login()
QTimer.singleShot(4000, stack.showMaximized)                                           ##using show function which is important without which everything is invisible
splash()
desktop_app.exec_()                                    

