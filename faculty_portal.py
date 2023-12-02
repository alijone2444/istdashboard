from PyQt5.QtWidgets import (QApplication, QMainWindow,QLabel,QPushButton,QMenu,QAction,QFrame,QScrollArea,QGroupBox,QFormLayout,QGridLayout,QVBoxLayout
,QVBoxLayout,QFormLayout,QAbstractItemView,QWidget,QTableWidget,QHBoxLayout,QTableWidgetItem,QStyledItemDelegate,QLineEdit,QFileDialog, QHeaderView)
from PyQt5.QtGui import QFont,QPixmap,QIcon,QMovie,QPainter, QPen,QColor
from PyQt5.QtCore import Qt
import sys
from reqest_and_response import RequestAndResponse
import pandas as pd
import openpyxl
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        # last column
        if index.column() == (index.model().columnCount() - 1 ):
            return super().createEditor(parent, option, index)

        if index.column() == (index.model().columnCount() - 2 ):
            return super().createEditor(parent, option, index)


class Faculty(QMainWindow):
    date = None
    maxmarks = None
    def __init__(self):
        super().__init__()
        w=1000
        h=640
        self.resize(w,h)
        self.setStyleSheet( '''QMainWindow{background-color:black}QScrollBar{ background-color: darkgrey}QTableWidget::item {background-color:lightblue;border: 1px solid black;color:black}
            QTableWidget::item::hover{background-color:white}QTableWidget::item::selected{background-color:white}
            QHeaderView::section {border:1px solid black;}QHeaderView:section:hover{background-color:lightblue}''')
        self.setWindowTitle("Faculty")
        self.setWindowIcon(QIcon('logo'))

        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.gridlayout =QGridLayout(self.widget)
        self.setLayout(self.gridlayout)

        self.grid3 = QGridLayout(self)
        self.label = QLabel("None",self)
        self.label.setStyleSheet("background-color:black;")
        self.label.setMaximumSize(1500,800)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setLayout(self.grid3)

        self.movie = QMovie("loader.gif")
        self.label.setMovie(self.movie)
        self.label.setScaledContents(True)
        self.movie.setSpeed(200)
        self.movie.start()

        self.hb = QHBoxLayout(self)
        self.empty_label2 = QLabel(self)
        self.empty_label2.setFixedSize(170,50)
        self.empty_label2.setStyleSheet("border:2px solid transparent")
        self.empty_label2.setLayout(self.hb)
        self.Faculty_label = QLabel(" Faculty",self)
        self.Faculty_label.setStyleSheet("background-color:transparent;color:rgba(72,221,253,255);border-style:none")
        self.Faculty_label.setFixedSize(150,30)
        self.Faculty_label.setFont(QFont("arial",15))
        self.goback = QPushButton("<",self)
        self.goback.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.goback.setFixedSize(30,30)
        self.goback.clicked.connect(lambda:portal.going_back_to_dashboard())
        self.hb.addWidget(self.goback)
        self.hb.addWidget(self.Faculty_label)


        self.add_attendance = QPushButton("Add Attendance",self)
        self.add_attendance.setFixedSize(150,30)
        self.add_attendance.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.custom = QAction("Add Attendance Customly", self)
        self.upload_excel= QAction("Upload Excel File", self)
        menu3 = QMenu()
        menu3.addAction(self.custom)
        menu3.addAction(self.upload_excel)
        self.add_attendance.setMenu(menu3)
        menu3.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255)}
            QMenu:pressed{background-color:lightblue}''')
        self.custom.triggered.connect(lambda:window.showeditabletable())
        self.upload_excel.triggered.connect(lambda:window.upload_excelfile())
        self.attendance_information = QPushButton("attendance information",self)
        self.attendance_information.setStyleSheet("background-color:black;color:rgba(72,221,253,255);border:2px solid rgba(72,221,253,255);border-radius :7px;border-radius :7px;")
        self.attendance_information.setFixedSize(150,30)
        self.attendance_information.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')
        self.assembly_language = QAction("computer organization and assembly lamguage")
        self.assembly_language_lab = QAction("computer organization and assembly lamguage lab")
        self.data_structure = QAction("data structure")
        self.data_structure_lab = QAction("data structure lab")
        self.discrete_structure = QAction("discrete structure")
        self.graph_theory = QAction("graph theory")
        self.profesional_practices = QAction("professional practices",print("profesional_practices"))       
        self.menu = QMenu()
        menu1 = self.menu.addMenu("courses")
        menu1.addAction(self.assembly_language)
        menu1.addAction(self.assembly_language_lab)
        menu1.addAction(self.data_structure)
        menu1.addAction(self.data_structure_lab)
        menu1.addAction(self.graph_theory)
        menu1.addAction(self.profesional_practices)
        menu1.addAction(self.assembly_language_lab)
        self.attendance_information.setMenu(self.menu)
        self.menu.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255)}
            QMenu:pressed{background-color:lightblue}''')
        self.assembly_language.triggered.connect(lambda:window.showtable()) 
        

        self.result = QPushButton("Marks upload",self)
        self.result.setFixedSize(150,30)
        self.result.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}
            ''')

        self.final= QAction("BS CS FINAL", self)
        self.oht1= QAction("OHT1", self)
        self.oht2= QAction("OHT2", self)
        self.quiz= QAction("Quiz", self)
        self.assignment= QAction("Assignments", self)
        self.project= QAction("Projects", self)
        self.presentation= QAction("Presentations", self)
        menu2 = QMenu()
        menu2.addAction(self.quiz)
        menu2.addAction(self.assignment)
        menu2.addAction(self.project)
        menu2.addAction(self.presentation)
        menu2.addAction(self.oht1)
        menu2.addAction(self.oht2)
        menu2.addAction(self.final)
        self.result.setMenu(menu2)
        menu2.setStyleSheet('''QMenu{background-color:rgba(72,221,253,255);
            border:2px solid rgba(72,221,253,255);border-top:2px solid black;}
            QMenu.item.selected {color:rgba(72,221,253,255)}
            QMenu:pressed{background-color:lightblue}''')
        print(self.result.text())
        self.quiz.triggered.connect(lambda:window.marks_entry())

        self.grade_criteria = QPushButton("set Marks Distribution",self)
        self.grade_criteria.setFixedSize(150,30)
        self.grade_criteria.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}''')
        self.grade_criteria.clicked.connect(lambda:window.showgradingcriteria())

        self.grade_generation = QPushButton("Grade Generation",self)
        self.grade_generation.setFixedSize(150,30)
        self.grade_generation.setStyleSheet('''QPushButton{background-color:black;color:rgba(72,221,253,255);border-radius:5px;border:2px solid #52bdf6;}
            QPushButton:hover{border:2px solid white;}
            QPushButton:pressed{background-color:grey}''')

        self.form_for_buttons = QFormLayout(self)
        self.empty_label_for_buttons = QLabel(self)
        self.empty_label_for_buttons.setStyleSheet("border:10px solid #52bdf6;border-style:double;border-right:transparent")
        self.empty_label_for_buttons.setFixedSize(220,700)
        self.empty_label_for_buttons.setLayout(self.form_for_buttons)
        self.form_for_buttons.addWidget(self.empty_label2)
        self.form_for_buttons.addWidget(self.add_attendance)
        self.form_for_buttons.addWidget(self.attendance_information)
        self.form_for_buttons.addWidget(self.result)
        self.form_for_buttons.addWidget(self.grade_criteria)
        self.form_for_buttons.addWidget(self.grade_generation)
        
        self.scroll = QScrollArea(self)
        self.scroll.setWidget(self.empty_label_for_buttons)
        self.scroll.setWidgetResizable(True)
        self.scroll.setMaximumSize(220,800)
        self.scroll.setStyleSheet("border:2px solid #52bdf6;background-color:transparent")

        self.gridlayout.addWidget(self.label,0,0,5,5)         
        self.gridlayout.addWidget(self.scroll,0,0,5,1)        
    def showeditabletable(self):
        self.empty_label1 = QLabel(self)
        self.empty_label1.setStyleSheet("color:black;background-color:white")
        self.empty_label1.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.upload = QPushButton("Upload",self)
        self.upload.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.upload.setFixedSize(120,30)
        self.upload.clicked.connect(lambda:upload())
        self.attendance_label = QLabel("Students Attendance",self)
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label1.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.upload)

        self.editable_tableWidget = QTableWidget(40,5,self)
        self.editable_tableWidget.setMaximumSize(1500,800)        
        self.editable_tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        maxrow = 50
        maxcol = 5
        self.editable_tableWidget.setRowCount(maxrow) 
        self.editable_tableWidget.setColumnCount(maxcol)  


        def QString(a):
            return (a)
        self.editable_tableWidget.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))
        self.gridlayout.addWidget(self.empty_label1,0,1,1,5)
        self.gridlayout.addWidget(self.editable_tableWidget,1,1,4,5)
        def upload():
            whole = {}
            x=""
            counter = -1
            for i in range(maxrow):
                individual_dict = {}
                for j in range(maxcol):

                    thing = self.editable_tableWidget.item(i,j)
                    if thing is not None and thing.text() != '':
                        items = (thing.text())
                        x=str(items)
                        if j%5==1:
                            individual_dict['date']=x
                        elif j%5==2:
                            individual_dict['name']=x
                        elif j%5==3:
                            individual_dict['regno']=int(x)
                        elif j%5==4:
                            individual_dict['status']=x
                            counter = counter+1
                            whole[counter] = individual_dict
            print(whole)            
            for i in whole:
                getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
                print(whole[i])
                getdata.post_data(whole[i])

                        
    def upload_excelfile(self):
        self.empty_label1 = QLabel(self)
        self.empty_label1.setStyleSheet("color:black;background-color:white")
        self.empty_label1.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.upload_excel = QPushButton("Upload Excel",self)
        self.upload_excel.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.upload_excel.setFixedSize(120,30)
        self.attendance_label = QLabel("Students Attendance",self)
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label1.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.upload_excel)

        self.excel_attendance = QTableWidget(40,5,self)
        self.excel_attendance.setMaximumSize(1599,800)
        maxrow = 50
        maxcol = 5
        self.excel_attendance.setRowCount(maxrow) 
        self.excel_attendance.setColumnCount(maxcol) 
        self.excel_attendance.verticalHeader().setVisible(False)
        self.excel_attendance.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        def dialog():
            file = QFileDialog(self)
            self.dialog = file.getOpenFileName(self, 'Open attendance', '', '(*.xlsx)')
            print(self.dialog[1])
            path = self.dialog[0]
            wb_obj = openpyxl.load_workbook(path) 
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            sr = []
            regno = []
            name = []
            status = []
            counter = 2
            for i in range(1,sheet_obj.max_column):
                for j in range(3,sheet_obj.max_row-1):
                    cell_obj = sheet_obj.cell(row =j, column = i)
                    print(i,cell_obj.value)
                    if i == 1:
                        sr.append(cell_obj.value)
                    elif i == 2:
                        regno.append(cell_obj.value)
                    elif i == 3:
                        name.append(cell_obj.value)
                    elif i == 4:
                        status.append(cell_obj.value)
            #print(sr,regno,name,status)
            for i in range(0,len(sr)):
                for j in range(5):
                    item = QTableWidgetItem()
                    if j ==0:
                        item.setData(Qt.EditRole, sr[i])
                        self.excel_attendance.setItem(i,j, item)
                    elif j==1:
                        item.setData(Qt.EditRole,"12/24/2022")
                        self.excel_attendance.setItem(i,j, item)
                    elif j==2:
                        self.excel_attendance.resizeColumnToContents(i)
                        item.setData(Qt.EditRole,name[i])
                        self.excel_attendance.setItem(i,j, item)
                    
                    elif j==3:
                        item.setData(Qt.EditRole,regno[i])
                        self.excel_attendance.setItem(i,j, item)
                    elif j==4:
                        item.setData(Qt.EditRole,status[i])
                        self.excel_attendance.setItem(i,j, item)
                    
                    print(i,j,sr[i])

            whole = {}
            x=""
            counter = -1
            for i in range(maxrow):
                individual_dict = {}
                for j in range(maxcol):
                    print(i,j)
                    thing = self.excel_attendance.item(i,j)
                    if thing is not None and thing.text() != '':
                        items = (thing.text())
                        x=str(items)
                        if j%5==1:
                            individual_dict['date']=x
                        elif j%5==2:
                            individual_dict['name']=x
                        elif j%5==3:
                            individual_dict['regno']=int(x)
                        elif j%5==4:
                            individual_dict['status']=x
                            counter = counter+1
                            whole[counter] = individual_dict
            print(whole)
            self.send = QPushButton("SEND(TO SERVER)",self)
            self.send.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
            self.send.setFixedSize(120,30)
            self.hvbox.addWidget(self.send)
            self.upload_excel.hide()
            def set_database():                
                print("ok")
                #for i in whole:
                 #   getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
                  #  print(whole[i])
                   # getdata.post_data(whole[i])
            self.send.clicked.connect(lambda:set_database())


        self.upload_excel.clicked.connect(lambda:dialog())

        def QString(a):
            return (a)
        self.excel_attendance.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))
        self.gridlayout.addWidget(self.empty_label1,0,1,1,5)
        self.gridlayout.addWidget(self.excel_attendance,1,1,4,5)
    def showtable(self):
        self.empty_label = QLabel(self)
        self.empty_label.setStyleSheet("color:black;background-color:white")
        self.empty_label.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.edit_button = QPushButton("EDIT",self)
        self.edit_button.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.edit_button.setFixedSize(120,30)
        self.attendance_label = QLabel("Students Attendance",self)
        self.attendance_label.setAlignment(Qt.AlignCenter)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.edit_button)

        self.tableWidget = QTableWidget(40,5,self) 
        self.tableWidget.setMaximumSize(1500,800)        
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.verticalHeader().setVisible(False)
      #  lastIndex = 18
     #   item = self.tableWidget.item(lastIndex, 0)
    #    self.tableWidget.scrollToItem(item, QAbstractItemView.PositionAtTop)
   #     self.tableWidget.selectRow(lastIndex)
  #      self.tableWidget.verticalHeader().setVisible(False)    
     #   self.tableWidget.
 #       delegate = ReadOnlyDelegate(self.tableWidget)
#        self.tableWidget.setItemDelegate(delegate)

        getdata = RequestAndResponse("http://127.0.0.1:8000/faculty attendance/")
        result = getdata.get_response()
        print(result)
        ids,date,name,rollno,status= [],[],[],[],[]
        for i in range (len(result)):
            ids.append(result[i]['id'])
            date.append(result[i]['date'])
            name.append(result[i]['name'])
            rollno.append(result[i]['regno'])
            status.append(result[i]['status'])
        maxrow = 50
        maxcol = 5
        self.tableWidget.setRowCount(maxrow) 
        self.tableWidget.setColumnCount(maxcol)  
        currentRowCount = -1
        for i in range(len(ids)):
            currentRowCount += 1
            print("currentRowCount",currentRowCount)
            self.tableWidget.setItem(currentRowCount, 0, QTableWidgetItem(str(i+1)))
            self.tableWidget.setItem(currentRowCount, 1, QTableWidgetItem(str(date[i])))
            self.tableWidget.setItem(currentRowCount, 2, QTableWidgetItem(str(name[i])))
            self.tableWidget.setItem(currentRowCount, 3, QTableWidgetItem(str(rollno[i])))
            self.tableWidget.setItem(currentRowCount, 4, QTableWidgetItem(str(status[i])))

        def QString(a):
            return (a)
        self.tableWidget.setHorizontalHeaderLabels(QString("Sr;Date;Name;Rollno;Status").split(";"))
        self.gridlayout.addWidget(self.empty_label,0,1,1,5)
        self.gridlayout.addWidget(self.tableWidget,1,1,4,5)
    

    def marks_entry(self):
        self.empty_label = QLabel(self)
        self.empty_label.setStyleSheet("color:black;background-color:white")
        self.empty_label.setMaximumSize(1500,50)
        self.hvbox = QHBoxLayout(self)
        self.done = QPushButton("Done",self)
        self.done.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
        self.done.setFixedSize(60,30)
        self.maxnum = QLabel("Max marks:",self)
        self.maxnum.setFixedSize(60,30)
        self.lineedit = QLineEdit(self)
        self.lineedit.setFixedSize(130,30)
        self.done.clicked.connect(lambda:afterdone())
     
        def afterdone():
            Faculty.date = self.lineedit2.text()
            Faculty.maxmarks = self.lineedit.text()
            self.lineedit.hide()
            self.lineedit2.hide()
            self.maxnum.hide()
            self.date.hide()
            self.done.hide()
            string = "Date : "+Faculty.date+"\t\t Maximum Marks of quiz :"+Faculty.maxmarks
            self.label_to_show_data = QLabel(self)
            self.label_to_show_data.setText(string)

            self.upload = QPushButton("upload",self)
            self.upload.setFixedSize(130,30)
            self.upload.setStyleSheet('''QPushButton{background-color:lightgrey;border-radius:10px}
            QPushButton:hover{border:2px solid black;}
            QPushButton:pressed{background-color:white}
            ''')
            self.hvbox.addWidget(self.upload)
            self.hvbox.addWidget(self.label_to_show_data)
            self.upload.clicked.connect(lambda:uploadmarks())
            
        self.date = QLabel("Date:",self)
        self.date.setFixedSize(50,30)
        self.lineedit2 = QLineEdit(self)
        self.lineedit2.setFixedSize(130,30)

        self.attendance_label = QLabel("Students Marks Entry",self)
        self.attendance_label.setFont(QFont("arial",12))
        self.empty_label.setLayout(self.hvbox)
        self.hvbox.addWidget(self.attendance_label)
        self.hvbox.addWidget(self.date)
        self.hvbox.addWidget(self.lineedit2)
        self.hvbox.addWidget(self.maxnum)
        self.hvbox.addWidget(self.lineedit)
        self.hvbox.addWidget(self.done)


        self.tableWidget1 = QTableWidget(40,4,self)
        self.tableWidget1.setMaximumSize(1500,800)        
        self.tableWidget1.verticalHeader().setVisible(False)
        self.tableWidget1.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)


        maxrow = 50
        maxcol = 4
        self.tableWidget1.setRowCount(maxrow) 
        self.tableWidget1.setColumnCount(maxcol)  


        def QString(a):
            return (a)
        self.tableWidget1.setHorizontalHeaderLabels(QString("Sr;Name;Rollno;Obtain Marks").split(";"))
        self.gridlayout.addWidget(self.empty_label,0,1,1,5)
        self.gridlayout.addWidget(self.tableWidget1,1,1,4,5)

        def uploadmarks():
            marks = {}
            x=""
            counter = -1
            for i in range(maxcol):
                individual_dict = {}
                for j in range(maxrow):

                    thing = self.tableWidget1.item(i,j)
                    if thing is not None and thing.text() != '':
                        items = (thing.text())
                        x=str(items)
                       # individual_dict['date'
                        print(j)
                        individual_dict['typeofexam'] = "Quiz"
                        individual_dict['date'] = str(Faculty.date)
                        if j%4==1:
                            individual_dict['name']=str(x)
                        elif j%4==2:
                            individual_dict['rollno']=int(x)
                        elif j%4==3:
                            individual_dict['maxnum']=int(Faculty.maxmarks)
                            individual_dict['obtainnum']=int(x)
                            counter = counter+1
                            marks[counter] = individual_dict

            print(marks)
            for i in marks:
                getdata = RequestAndResponse("http://127.0.0.1:8000/quiz marks/")
                print(marks[i])
                getdata.post_data(marks[i])

    def showgradingcriteria(self):
        self.empty_label_for_grade = QLabel(self)
        self.empty_label_for_grade.setMaximumSize(1500,800)        
        self.empty_label_for_grade.setStyleSheet("background-color:white")
        self.gridlayout.addWidget(self.empty_label_for_grade,0,1,5,5)
        self.nestedgrid = QGridLayout(self)
        self.empty_label_for_grade.setLayout(self.nestedgrid)
        self.setcriteria = QLabel("𝐌𝐀𝐑𝐊𝐒 𝐃𝐢𝐬𝐭𝐫𝐢𝐛𝐮𝐭𝐢𝐨𝐧",self)
        self.setcriteria.setFont(QFont("arial",15))
        self.setcriteria.setAlignment(Qt.AlignCenter)

        self.vbox1 = QVBoxLayout(self)
        self.nestedlabel = QLabel(self)
        self.nestedlabel.setLayout(self.vbox1)
        self.quizlabel = QLabel("ADD QUIZ PERCENTAGE",self)
        self.lineedit_for_quiz = QLineEdit("0",self)
        self.lineedit_for_quiz.setFixedSize(150,30)      
        
        self.assignment = QLabel("ADD ASSIGNMENT PERCENTAGE",self)
        self.lineedit_for_assignment = QLineEdit("0",self)
        self.lineedit_for_assignment.setFixedSize(150,30)

        self.project = QLabel("ADD PROJECT PERCENTAGE",self)
        self.lineedit_for_project = QLineEdit("0",self)
        self.lineedit_for_project.setFixedSize(150,30)

        self.ohtlabel = QLabel("ADD OHT'S PERCENTAGE",self)
        self.lineedit_for_oht = QLineEdit("0",self)
        self.lineedit_for_oht.setFixedSize(150,30)

        self.final = QLabel("ADD Final PERCENTAGE",self)
        self.lineedit_for_final = QLineEdit("0",self)
        self.lineedit_for_final.setFixedSize(150,30)

        self.set = QPushButton("SET Marks Distribution",self)
        self.set.setStyleSheet('''QPushButton{background-color:lightgrey;color:black;border-radius:5px;border:2px solid grey;}
            QPushButton:hover{border:2px solid black;background-color:darkgrey}
            QPushButton:pressed{background-color:grey}''')
        self.set.setFixedSize(140,30)

        self.update = QLabel("TOTAL MUST BE = 100%",self)
        self.update.setFixedSize(150,160)
        self.update.setAlignment(Qt.AlignCenter)
        self.update.setFont(QFont("arial",10))


        self.vbox1.addStretch(1) 
        self.vbox1.addWidget(self.quizlabel)
        self.vbox1.addWidget(self.lineedit_for_quiz)
        self.vbox1.addWidget(self.assignment)
        self.vbox1.addWidget(self.lineedit_for_assignment)
        self.vbox1.addWidget(self.project)
        self.vbox1.addWidget(self.lineedit_for_project)
        self.vbox1.addWidget(self.ohtlabel)
        self.vbox1.addWidget(self.lineedit_for_oht)
        self.vbox1.addWidget(self.final)
        self.vbox1.addWidget(self.update)
        self.vbox1.addWidget(self.lineedit_for_final)
        self.vbox1.addWidget(self.update) 
        self.vbox1.addWidget(self.set)

        q =1
        a=1
        o=1
        f=1
        p=1
        series = QPieSeries()
        series.append("final", 1).setColor(QColor("skyblue"))
        series.append(" Assignments", 1).setColor(QColor("lightblue"))
        series.append("OHT's", 1).setColor(QColor("darkblue"))
        series.append("Project",1).setColor(QColor("green"))
        series.append("quiz", 1).setColor(QColor("lightgreen"))

        slice = QPieSlice()
        slice = series.slices()[0]
        slice.setExploded(True)
        slice.setLabelVisible(True)
        slice.setPen(QPen(Qt.red, 2))
        slice.setBrush(Qt.green)

 
        chart = QChart()
        chart.addSeries(series)
        chart.createDefaultAxes()
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setTitle("CHART") 
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)
        chartview = QChartView(chart)
        chartview.setRenderHint(QPainter.Antialiasing)
        #chartview.setMaximumSize(600,280)
        chart.createDefaultAxes()


        self.nestedgrid.addWidget(self.setcriteria,0,0,1,5)
        self.nestedgrid.addWidget(self.nestedlabel,1,0,4,1)
        self.nestedgrid.addWidget(chartview,1,1,1,1)

        def getpercentage():
            self.update.setWordWrap(True)
            answer = 0
            try:
                answer = int(self.lineedit_for_quiz.text()) + int(self.lineedit_for_assignment.text()) + int(self.lineedit_for_project.text()) + int(self.lineedit_for_oht.text()) + int(self.lineedit_for_final.text())
            except:
                self.update.setText("ENTER INT(PERCENTAGE VALUE)")
            if answer>100:
                self.update.setText("EXCEEDING 100%")
            elif answer<100:
                self.update.setText("BELOW 100%")
            else:
                print(answer)
                q = int(self.lineedit_for_quiz.text())
                a = int(self.lineedit_for_assignment.text())
                p = int(self.lineedit_for_project.text())
                o = int(self.lineedit_for_oht.text())
                f = int(self.lineedit_for_final.text())
                series.clear()
                series2 = QPieSeries()
                series2.append("Quiz", q).setColor(QColor("skyblue"))
                series2.append(" Assignments", a).setColor(QColor("lightblue"))
                series2.append("OHT's", o).setColor(QColor(QColor("darkblue")))
                series2.append("final", f).setColor(QColor("lightgreen"))
                series2.append("Project",p).setColor(QColor("green"))
                chart.addSeries(series2)

                slice = QPieSlice()
                input_list = []
                input_list.append(q)
                input_list.append(a)
                input_list.append(p)
                input_list.append(o)
                input_list.append(f)
                max_value = max(input_list)
                index = input_list.index(max_value)
                print(index)
                slice = series2.slices()[index]
                slice.setExploded(True)
                slice.setLabelVisible(True)
                slice.setPen(QPen(Qt.red, 2))
                slice.setBrush(Qt.green)
        self.set.clicked.connect(lambda:getpercentage())
app = QApplication(sys.argv)
window = Faculty()
window.show()  
app.exec()